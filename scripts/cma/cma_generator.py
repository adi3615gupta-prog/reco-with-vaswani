#!/usr/bin/env python3
import sys
import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# -------------------------------------------------------------------------
# CONSTANTS & STYLE SYSTEM (Finance Navy & Teal Palette)
# -------------------------------------------------------------------------
FONT_FAMILY = "Segoe UI"
COLOR_PRIMARY_NAVY = "1F4E79"  # Header fill
COLOR_ACCENT_TEAL = "E2EFDA"   # Highlight / Total fill
COLOR_ZEBRA_LIGHT = "F2F2F2"   # Alternate rows
COLOR_WHITE = "FFFFFF"
COLOR_GRAY_BORDER = "D3D3D3"

# Fonts
font_title = Font(name=FONT_FAMILY, size=16, bold=True, color="1F4E79")
font_section = Font(name=FONT_FAMILY, size=12, bold=True, color="1F4E79")
font_header = Font(name=FONT_FAMILY, size=11, bold=True, color=COLOR_WHITE)
font_bold_data = Font(name=FONT_FAMILY, size=10, bold=True)
font_regular_data = Font(name=FONT_FAMILY, size=10)
font_italic_sub = Font(name=FONT_FAMILY, size=9, italic=True, color="595959")
font_recommendation = Font(name=FONT_FAMILY, size=10, bold=True, color="1F4E79")

# Fills
fill_header = PatternFill(start_color=COLOR_PRIMARY_NAVY, end_color=COLOR_PRIMARY_NAVY, fill_type="solid")
fill_total = PatternFill(start_color=COLOR_ACCENT_TEAL, end_color=COLOR_ACCENT_TEAL, fill_type="solid")
fill_zebra = PatternFill(start_color=COLOR_ZEBRA_LIGHT, end_color=COLOR_ZEBRA_LIGHT, fill_type="solid")
fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Borders
border_thin = Border(
    left=Side(style='thin', color=COLOR_GRAY_BORDER),
    right=Side(style='thin', color=COLOR_GRAY_BORDER),
    top=Side(style='thin', color=COLOR_GRAY_BORDER),
    bottom=Side(style='thin', color=COLOR_GRAY_BORDER)
)
border_header = Border(
    left=Side(style='thin', color="FFFFFF"),
    right=Side(style='thin', color="FFFFFF"),
    top=Side(style='medium', color=COLOR_PRIMARY_NAVY),
    bottom=Side(style='medium', color=COLOR_PRIMARY_NAVY)
)
border_total = Border(
    top=Side(style='thin', color="000000"),
    bottom=Side(style='double', color="000000")
)

# Alignments
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_center = Alignment(horizontal='center', vertical='center')
align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Number formats
FMT_CURRENCY = '₹#,##,##0.00'
FMT_PERCENT = '0.0%'
FMT_INTEGER = '#,##0'

# -------------------------------------------------------------------------
# DEFAULT PAYLOAD DATA (Fallback if no JSON input is supplied)
# -------------------------------------------------------------------------
DEFAULT_PAYLOAD = {
    "client_metadata": {
        "company_name": "Astra Tech Manufacturing Private Limited",
        "cin_number": "U29100MH2024PTC412345",
        "registered_address": "Plot No. 45, MIDC Industrial Area, Andheri East, Mumbai - 400093",
        "audit_firm_name": "Vaswani & Associates, Chartered Accountants",
        "firm_reg_no": "123456W",
        "partner_name": "CA Anish Vaswani",
        "membership_no": "089456",
        "udin": "26089456AAAAAB1234",
        "director_1_name": "Siddharth Malhotra",
        "director_2_name": "Neha Sharma"
    },
    "assumptions": {
        "projection_years": [2026, 2027, 2028, 2029, 2030],
        "capacity_utilization_pct": [60.0, 70.0, 80.0, 90.0, 95.0],
        "sales_growth_pct": [10.0, 15.0, 12.0, 10.0, 8.0],
        "rm_cost_pct": [55.0, 54.0, 53.5, 53.0, 53.0],
        "other_expenses_pct": [12.0, 11.5, 11.0, 10.5, 10.5],
        "tax_rate_pct": 25.17,
        "depreciation_method": "WDV",
        "depreciation_rates": {
            "building": 10.0,
            "plant_machinery": 15.0,
            "computers": 40.0,
            "office_equipment": 10.0
        }
    },
    "financial_inputs": {
        "base_year_sales": 45000000.0,
        "share_capital_init": 12000000.0,
        "fixed_assets_init": {
            "building": 15000000.0,
            "plant_machinery": 22000000.0,
            "computers": 1500000.0,
            "office_equipment": 2500000.0
        },
        "capex_plan": [
            {
                "year": 1,
                "building": { "gt_180_days": 2000000.0, "lt_180_days": 500000.0 },
                "plant_machinery": { "gt_180_days": 4000000.0, "lt_180_days": 1000000.0 },
                "computers": { "gt_180_days": 200000.0, "lt_180_days": 50000.0 },
                "office_equipment": { "gt_180_days": 300000.0, "lt_180_days": 100000.0 }
            },
            {
                "year": 2,
                "building": { "gt_180_days": 0.0, "lt_180_days": 0.0 },
                "plant_machinery": { "gt_180_days": 1500000.0, "lt_180_days": 0.0 },
                "computers": { "gt_180_days": 100000.0, "lt_180_days": 0.0 },
                "office_equipment": { "gt_180_days": 200000.0, "lt_180_days": 0.0 }
            },
            {
                "year": 3,
                "building": { "gt_180_days": 0.0, "lt_180_days": 0.0 },
                "plant_machinery": { "gt_180_days": 0.0, "lt_180_days": 0.0 },
                "computers": { "gt_180_days": 0.0, "lt_180_days": 0.0 },
                "office_equipment": { "gt_180_days": 0.0, "lt_180_days": 0.0 }
            },
            {
                "year": 4,
                "building": { "gt_180_days": 0.0, "lt_180_days": 0.0 },
                "plant_machinery": { "gt_180_days": 1000000.0, "lt_180_days": 0.0 },
                "computers": { "gt_180_days": 50000.0, "lt_180_days": 0.0 },
                "office_equipment": { "gt_180_days": 0.0, "lt_180_days": 0.0 }
            },
            {
                "year": 5,
                "building": { "gt_180_days": 0.0, "lt_180_days": 0.0 },
                "plant_machinery": { "gt_180_days": 0.0, "lt_180_days": 0.0 },
                "computers": { "gt_180_days": 0.0, "lt_180_days": 0.0 },
                "office_equipment": { "gt_180_days": 0.0, "lt_180_days": 0.0 }
            }
        ]
    },
    "loan_details": {
        "term_loan_amount": 15000000.0,
        "interest_rate_pct": 10.5,
        "repayment_months": 60,
        "moratorium_months": 6,
        "cc_limit": 8000000.0
    },
    "drawing_power_inputs": {
        "stock_value": 11000000.0,
        "sundry_creditors": 3500000.0,
        "debtors_under_90_days": 85000000.0,
        "debtors_over_90_days": 15000000.0,
        "stock_margin_pct": 25.0,
        "debtors_margin_pct": 40.0
    }
}

# -------------------------------------------------------------------------
# EXCEL STYLING AND BUILD HELPER FUNCTIONS
# -------------------------------------------------------------------------
def apply_corporate_header(ws, start_row, title, subtitle=None):
    ws.row_dimensions[start_row].height = 28
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = font_title
    cell.alignment = align_left
    
    if subtitle:
        ws.row_dimensions[start_row + 1].height = 18
        sub_cell = ws.cell(row=start_row + 1, column=1, value=subtitle)
        sub_cell.font = font_italic_sub
        sub_cell.alignment = align_left
        return 3
    return 2

def style_row_range(ws, row_idx, col_start, col_end, font=None, fill=None, alignment=None, border=None, num_format=None):
    for col_idx in range(col_start, col_end + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if num_format:
            cell.number_format = num_format

def autofit_columns(ws, padding=3, min_width=10):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value or '')
            if val.startswith('='):
                # Simple heuristic: don't count formula length as cell text length
                val = "123,456.00"
            max_len = max(max_len, len(val))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + padding, min_width)

# -------------------------------------------------------------------------
# SHEET GENERATORS
# -------------------------------------------------------------------------

def build_assumptions_sheet(ws, data):
    ws.title = "Assumptions"
    ws.views.sheetView[0].showGridLines = True
    
    # 1. Header
    r = apply_corporate_header(ws, 1, "PROJECT FINANCIAL ASSUMPTIONS", "CMA & Loan Sanction Sensitivity Baseline")
    
    # Table 1: Macro Projections
    ws.cell(row=r, column=1, value="Core Assumptions & Growth Driver").font = font_section
    r += 1
    
    headers = ["Parameter", "Year 1 (Y1)", "Year 2 (Y2)", "Year 3 (Y3)", "Year 4 (Y4)", "Year 5 (Y5)"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # Fill Table 1
    ass = data["assumptions"]
    fin = data["financial_inputs"]
    loan = data["loan_details"]
    
    rows_data = [
        ("Capacity Utilization %", [v / 100.0 for v in ass["capacity_utilization_pct"]], FMT_PERCENT),
        ("Sales Growth % (YoY)", [v / 100.0 for v in ass["sales_growth_pct"]], FMT_PERCENT),
        ("Raw Material Cost % of Sales", [v / 100.0 for v in ass["rm_cost_pct"]], FMT_PERCENT),
        ("Other Operating Expenses % of Sales", [v / 100.0 for v in ass["other_expenses_pct"]], FMT_PERCENT),
    ]
    
    for label, vals, fmt in rows_data:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=label).font = font_bold_data
        ws.cell(row=r, column=1).alignment = align_left
        ws.cell(row=r, column=1).border = border_thin
        
        for idx, val in enumerate(vals):
            c = ws.cell(row=r, column=idx+2, value=val)
            c.font = font_regular_data
            c.number_format = fmt
            c.alignment = align_right
            c.border = border_thin
        r += 2  # Leaving a row gap between elements
        
    # Table 2: Loan Details
    r += 1
    ws.cell(row=r, column=1, value="Loan Sanction Parameters").font = font_section
    r += 1
    
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(["Sanction Parameter", "Assumed Baseline Value"]):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    loan_rows = [
        ("Term Loan Sanctioned (₹)", loan["term_loan_amount"], FMT_CURRENCY),
        ("CC Limit Sanctioned (₹)", loan["cc_limit"], FMT_CURRENCY),
        ("Term Loan Interest Rate (per Annum)", loan["interest_rate_pct"] / 100.0, FMT_PERCENT),
        ("Cash Credit Interest Rate (per Annum)", (loan["interest_rate_pct"] + 1.0) / 100.0, FMT_PERCENT), # CC interest is usually base+1
        ("Repayment Term (Months)", loan["repayment_months"], FMT_INTEGER),
        ("Moratorium Period (Months)", loan["moratorium_months"], FMT_INTEGER),
        ("Corporate Income Tax Rate", ass["tax_rate_pct"] / 100.0, FMT_PERCENT),
        ("CC Stock Margin %", data["drawing_power_inputs"]["stock_margin_pct"] / 100.0, FMT_PERCENT),
        ("CC Debtors Margin %", data["drawing_power_inputs"]["debtors_margin_pct"] / 100.0, FMT_PERCENT)
    ]
    
    for label, val, fmt in loan_rows:
        ws.row_dimensions[r].height = 20
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = font_bold_data
        c1.alignment = align_left
        c1.border = border_thin
        
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font = font_regular_data
        c2.number_format = fmt
        c2.alignment = align_right
        c2.border = border_thin
        r += 1
        
    # Table 3: Fixed Assets Baseline & Rates
    r += 2
    ws.cell(row=r, column=1, value="Fixed Asset Blocks & Depreciation Rates (WDV)").font = font_section
    r += 1
    
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(["Asset Block Class", "Starting Net Block Value (₹)", "Depreciation Rate (WDV)"]):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    asset_rows = [
        ("Building Block", fin["fixed_assets_init"]["building"], ass["depreciation_rates"]["building"] / 100.0),
        ("Plant & Machinery Block", fin["fixed_assets_init"]["plant_machinery"], ass["depreciation_rates"]["plant_machinery"] / 100.0),
        ("Computers & IT Block", fin["fixed_assets_init"]["computers"], ass["depreciation_rates"]["computers"] / 100.0),
        ("Office Equipment Block", fin["fixed_assets_init"]["office_equipment"], ass["depreciation_rates"]["office_equipment"] / 100.0)
    ]
    
    for label, init_val, dep_rate in asset_rows:
        ws.row_dimensions[r].height = 20
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = font_bold_data
        c1.border = border_thin
        
        c2 = ws.cell(row=r, column=2, value=init_val)
        c2.font = font_regular_data
        c2.number_format = FMT_CURRENCY
        c2.border = border_thin
        c2.alignment = align_right
        
        c3 = ws.cell(row=r, column=3, value=dep_rate)
        c3.font = font_regular_data
        c3.number_format = FMT_PERCENT
        c3.border = border_thin
        c3.alignment = align_right
        r += 1
        
    # Table 4: Base Year Historicals
    r += 2
    ws.cell(row=r, column=1, value="Preceding Year (Base) Financials").font = font_section
    r += 1
    ws.row_dimensions[r].height = 20
    
    ws.cell(row=r, column=1, value="Base Year Revenue (₹)").font = font_bold_data
    ws.cell(row=r, column=1).border = border_thin
    c_rev = ws.cell(row=r, column=2, value=fin["base_year_sales"])
    c_rev.font = font_regular_data
    c_rev.number_format = FMT_CURRENCY
    c_rev.alignment = align_right
    c_rev.border = border_thin
    r += 1
    
    ws.cell(row=r, column=1, value="Initial Share Capital (₹)").font = font_bold_data
    ws.cell(row=r, column=1).border = border_thin
    c_cap = ws.cell(row=r, column=2, value=fin["share_capital_init"])
    c_cap.font = font_regular_data
    c_cap.number_format = FMT_CURRENCY
    c_cap.alignment = align_right
    c_cap.border = border_thin
    r += 1
    
    ws.cell(row=r, column=1, value="Initial Cash Balance (₹)").font = font_bold_data
    ws.cell(row=r, column=1).border = border_thin
    c_cash = ws.cell(row=r, column=2, value=fin["base_year_sales"] * 0.015) # Assume 1.5% of sales
    c_cash.font = font_regular_data
    c_cash.number_format = FMT_CURRENCY
    c_cash.alignment = align_right
    c_cash.border = border_thin
    r += 1
    
    autofit_columns(ws)


def build_repayment_sheet(ws, data):
    ws.title = "Repayment"
    ws.views.sheetView[0].showGridLines = True
    
    r = apply_corporate_header(ws, 1, "TERM LOAN REPAYMENT SCHEDULE", "Monthly Amortization and Yearly Aggregations")
    
    # 1. Monthly Table on the left (cols A-F)
    ws.cell(row=r, column=1, value="Monthly Loan Amortization Schedule").font = font_section
    r += 1
    
    m_headers = ["Month", "Opening Balance", "Interest Accrued", "Principal Repayment", "Total EMI", "Closing Balance"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(m_headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    
    monthly_start_row = r + 1
    r += 1
    
    # Loop over 60 months
    total_months = data["loan_details"]["repayment_months"]
    for m in range(1, total_months + 1):
        ws.row_dimensions[r].height = 18
        
        ws.cell(row=r, column=1, value=m).alignment = align_center
        ws.cell(row=r, column=1).font = font_regular_data
        ws.cell(row=r, column=1).border = border_thin
        
        # Opening balance
        if m == 1:
            ws.cell(row=r, column=2, value="=Assumptions!$B$16").number_format = FMT_CURRENCY
        else:
            ws.cell(row=r, column=2, value=f"=F{r-1}").number_format = FMT_CURRENCY
        ws.cell(row=r, column=2).alignment = align_right
        ws.cell(row=r, column=2).font = font_regular_data
        ws.cell(row=r, column=2).border = border_thin
        
        # Interest: OpBal * Rate / 12
        ws.cell(row=r, column=3, value=f"=B{r}*(Assumptions!$B$18/12)").number_format = FMT_CURRENCY
        ws.cell(row=r, column=3).alignment = align_right
        ws.cell(row=r, column=3).font = font_regular_data
        ws.cell(row=r, column=3).border = border_thin
        
        # Principal: IF(m <= Moratorium, 0, SanctionAmount / (Term - Moratorium))
        ws.cell(row=r, column=4, value=f"=IF(A{r}<=Assumptions!$B$21,0,Assumptions!$B$16/(Assumptions!$B$20-Assumptions!$B$21))").number_format = FMT_CURRENCY
        ws.cell(row=r, column=4).alignment = align_right
        ws.cell(row=r, column=4).font = font_regular_data
        ws.cell(row=r, column=4).border = border_thin
        
        # EMI = Interest + Principal
        ws.cell(row=r, column=5, value=f"=C{r}+D{r}").number_format = FMT_CURRENCY
        ws.cell(row=r, column=5).alignment = align_right
        ws.cell(row=r, column=5).font = font_regular_data
        ws.cell(row=r, column=5).border = border_thin
        
        # Closing = Opening - Principal
        ws.cell(row=r, column=6, value=f"=B{r}-D{r}").number_format = FMT_CURRENCY
        ws.cell(row=r, column=6).alignment = align_right
        ws.cell(row=r, column=6).font = font_regular_data
        ws.cell(row=r, column=6).border = border_thin
        r += 1
        
    monthly_end_row = r - 1
    
    # 2. Yearly Amortization Summary (cols H-M)
    ws.cell(row=4, column=8, value="Yearly Summary Schedule").font = font_section
    y_headers = ["Year", "Opening Balance", "Interest Sum", "Principal Sum", "Total Debt Service", "Closing Balance"]
    
    ws.row_dimensions[5].height = 24
    for idx, h in enumerate(y_headers):
        cell = ws.cell(row=5, column=idx+8, value=h)
        cell.font = font_header
        cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        cell.alignment = align_header
        cell.border = border_header
        
    for y in range(1, 6):
        row_y = y + 5
        ws.row_dimensions[row_y].height = 20
        
        # Year
        ws.cell(row=row_y, column=8, value=y).alignment = align_center
        ws.cell(row=row_y, column=8).font = font_bold_data
        ws.cell(row=row_y, column=8).border = border_thin
        
        # Opening Bal (Month 1, Month 13, Month 25, Month 37, Month 49)
        m_start_idx = (y - 1) * 12 + monthly_start_row
        ws.cell(row=row_y, column=9, value=f"=B{m_start_idx}").number_format = FMT_CURRENCY
        ws.cell(row=row_y, column=9).alignment = align_right
        ws.cell(row=row_y, column=9).font = font_regular_data
        ws.cell(row=row_y, column=9).border = border_thin
        
        # Interest Sum (sum of 12 months)
        m_end_idx = m_start_idx + 11
        ws.cell(row=row_y, column=10, value=f"=SUM(C{m_start_idx}:C{m_end_idx})").number_format = FMT_CURRENCY
        ws.cell(row=row_y, column=10).alignment = align_right
        ws.cell(row=row_y, column=10).font = font_regular_data
        ws.cell(row=row_y, column=10).border = border_thin
        
        # Principal Sum (sum of 12 months)
        ws.cell(row=row_y, column=11, value=f"=SUM(D{m_start_idx}:D{m_end_idx})").number_format = FMT_CURRENCY
        ws.cell(row=row_y, column=11).alignment = align_right
        ws.cell(row=row_y, column=11).font = font_regular_data
        ws.cell(row=row_y, column=11).border = border_thin
        
        # Total Debt Service
        ws.cell(row=row_y, column=12, value=f"=J{row_y}+K{row_y}").number_format = FMT_CURRENCY
        ws.cell(row=row_y, column=12).alignment = align_right
        ws.cell(row=row_y, column=12).font = font_bold_data
        ws.cell(row=row_y, column=12).border = border_thin
        
        # Closing Bal (Month 12, Month 24, Month 36, Month 48, Month 60)
        ws.cell(row=row_y, column=13, value=f"=F{m_end_idx}").number_format = FMT_CURRENCY
        ws.cell(row=row_y, column=13).alignment = align_right
        ws.cell(row=row_y, column=13).font = font_regular_data
        ws.cell(row=row_y, column=13).border = border_thin

    autofit_columns(ws)


def build_depreciation_sheet(ws, data):
    ws.title = "Depreciation"
    ws.views.sheetView[0].showGridLines = True
    
    r = apply_corporate_header(ws, 1, "FIXED ASSETS & DEPRECIATION SCHEDULE", "WDV Capital Allowance and Asset Additions Registry")
    
    blocks = [
        ("Building Block", "Assumptions!$C$29", data["financial_inputs"]["fixed_assets_init"]["building"], "building"),
        ("Plant & Machinery Block", "Assumptions!$C$30", data["financial_inputs"]["fixed_assets_init"]["plant_machinery"], "plant_machinery"),
        ("Computers & IT Block", "Assumptions!$C$31", data["financial_inputs"]["fixed_assets_init"]["computers"], "computers"),
        ("Office Equipment Block", "Assumptions!$C$32", data["financial_inputs"]["fixed_assets_init"]["office_equipment"], "office_equipment")
    ]
    
    headers = ["Year", "Opening WDV", "Additions (>180d)", "Additions (<180d)", "Deletions", "Gross Block", "Depreciation", "Closing Net Block"]
    capex_data = data["financial_inputs"]["capex_plan"]
    
    block_row_refs = {} # To keep track of closing net block row indices for each block
    
    for block_name, rate_ref, init_val, key in blocks:
        ws.cell(row=r, column=1, value=block_name).font = font_section
        r += 1
        
        ws.row_dimensions[r].height = 24
        for idx, h in enumerate(headers):
            cell = ws.cell(row=r, column=idx+1, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_header
            cell.border = border_header
        r += 1
        
        block_start_row = r
        for y in range(1, 6):
            ws.row_dimensions[r].height = 20
            
            # Year
            ws.cell(row=r, column=1, value=y).alignment = align_center
            ws.cell(row=r, column=1).font = font_bold_data
            ws.cell(row=r, column=1).border = border_thin
            
            # Opening WDV
            if y == 1:
                ws.cell(row=r, column=2, value=init_val).number_format = FMT_CURRENCY
            else:
                ws.cell(row=r, column=2, value=f"=H{r-1}").number_format = FMT_CURRENCY
            ws.cell(row=r, column=2).alignment = align_right
            ws.cell(row=r, column=2).font = font_regular_data
            ws.cell(row=r, column=2).border = border_thin
            
            # Additions (>180)
            ws.cell(row=r, column=3, value=capex_data[y-1][key]["gt_180_days"]).number_format = FMT_CURRENCY
            ws.cell(row=r, column=3).alignment = align_right
            ws.cell(row=r, column=3).font = font_regular_data
            ws.cell(row=r, column=3).border = border_thin
            
            # Additions (<180)
            ws.cell(row=r, column=4, value=capex_data[y-1][key]["lt_180_days"]).number_format = FMT_CURRENCY
            ws.cell(row=r, column=4).alignment = align_right
            ws.cell(row=r, column=4).font = font_regular_data
            ws.cell(row=r, column=4).border = border_thin
            
            # Deletions
            ws.cell(row=r, column=5, value=0.0).number_format = FMT_CURRENCY
            ws.cell(row=r, column=5).alignment = align_right
            ws.cell(row=r, column=5).font = font_regular_data
            ws.cell(row=r, column=5).border = border_thin
            
            # Gross Block = Op + Add_gt + Add_lt - Del
            ws.cell(row=r, column=6, value=f"=B{r}+C{r}+D{r}-E{r}").number_format = FMT_CURRENCY
            ws.cell(row=r, column=6).alignment = align_right
            ws.cell(row=r, column=6).font = font_regular_data
            ws.cell(row=r, column=6).border = border_thin
            
            # Depreciation = (Op + Add_gt - Del) * Rate + Add_lt * Rate * 0.5
            ws.cell(row=r, column=7, value=f"=(B{r}+C{r}-E{r})*{rate_ref}+D{r}*{rate_ref}*0.5").number_format = FMT_CURRENCY
            ws.cell(row=r, column=7).alignment = align_right
            ws.cell(row=r, column=7).font = font_regular_data
            ws.cell(row=r, column=7).border = border_thin
            
            # Closing Net Block = Gross - Depreciation
            ws.cell(row=r, column=8, value=f"=F{r}-G{r}").number_format = FMT_CURRENCY
            ws.cell(row=r, column=8).alignment = align_right
            ws.cell(row=r, column=8).font = font_bold_data
            ws.cell(row=r, column=8).border = border_thin
            r += 1
            
        block_row_refs[key] = list(range(block_start_row, r))
        r += 2  # Gap between blocks
        
    # 5. Consolidated Summary Table at the bottom
    ws.cell(row=r, column=1, value="CONSOLIDATED FIXED ASSETS SUMMARY").font = font_section
    r += 1
    
    ws.row_dimensions[r].height = 24
    summary_headers = ["Year", "Opening Net Block", "Total Additions", "Total Deletions", "Gross Consolidated Block", "Total Depreciation", "Net Closing Block"]
    for idx, h in enumerate(summary_headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = PatternFill(start_color="16365C", end_color="16365C", fill_type="solid")
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    consolidated_start_row = r
    for y in range(1, 6):
        ws.row_dimensions[r].height = 20
        idx = y - 1
        
        # Year
        ws.cell(row=r, column=1, value=y).alignment = align_center
        ws.cell(row=r, column=1).font = font_bold_data
        ws.cell(row=r, column=1).border = border_thin
        
        # Helper lists of rows for each class for this year
        b_r = block_row_refs["building"][idx]
        pm_r = block_row_refs["plant_machinery"][idx]
        cp_r = block_row_refs["computers"][idx]
        oe_r = block_row_refs["office_equipment"][idx]
        
        # Opening Net Block Sum
        ws.cell(row=r, column=2, value=f"=B{b_r}+B{pm_r}+B{cp_r}+B{oe_r}").number_format = FMT_CURRENCY
        ws.cell(row=r, column=2).alignment = align_right
        ws.cell(row=r, column=2).font = font_regular_data
        ws.cell(row=r, column=2).border = border_thin
        
        # Total Additions (gt + lt for all classes)
        ws.cell(row=r, column=3, value=f"=C{b_r}+D{b_r}+C{pm_r}+D{pm_r}+C{cp_r}+D{cp_r}+C{oe_r}+D{oe_r}").number_format = FMT_CURRENCY
        ws.cell(row=r, column=3).alignment = align_right
        ws.cell(row=r, column=3).font = font_regular_data
        ws.cell(row=r, column=3).border = border_thin
        
        # Total Deletions
        ws.cell(row=r, column=4, value=f"=E{b_r}+E{pm_r}+E{cp_r}+E{oe_r}").number_format = FMT_CURRENCY
        ws.cell(row=r, column=4).alignment = align_right
        ws.cell(row=r, column=4).font = font_regular_data
        ws.cell(row=r, column=4).border = border_thin
        
        # Gross Consolidated Block
        ws.cell(row=r, column=5, value=f"=F{b_r}+F{pm_r}+F{cp_r}+F{oe_r}").number_format = FMT_CURRENCY
        ws.cell(row=r, column=5).alignment = align_right
        ws.cell(row=r, column=5).font = font_regular_data
        ws.cell(row=r, column=5).border = border_thin
        
        # Total Depreciation
        ws.cell(row=r, column=6, value=f"=G{b_r}+G{pm_r}+G{cp_r}+G{oe_r}").number_format = FMT_CURRENCY
        ws.cell(row=r, column=6).alignment = align_right
        ws.cell(row=r, column=6).font = font_bold_data
        ws.cell(row=r, column=6).border = border_thin
        
        # Net Closing Block
        ws.cell(row=r, column=7, value=f"=H{b_r}+H{pm_r}+H{cp_r}+H{oe_r}").number_format = FMT_CURRENCY
        ws.cell(row=r, column=7).alignment = align_right
        ws.cell(row=r, column=7).font = font_bold_data
        ws.cell(row=r, column=7).border = border_thin
        
        # Highlight total row
        style_row_range(ws, r, 1, 7, fill=fill_total)
        r += 1

    autofit_columns(ws)


def build_pnl_sheet(ws, data):
    ws.title = "Projected P&L"
    ws.views.sheetView[0].showGridLines = True
    
    r = apply_corporate_header(ws, 1, "PROJECTED STATEMENT OF PROFIT & LOSS (₹)", "5-Year Profitability Forecast & Margin Trajectory")
    r += 1
    
    headers = ["Particulars", "Base Year (Y0)", "Year 1 (Y1)", "Year 2 (Y2)", "Year 3 (Y3)", "Year 4 (Y4)", "Year 5 (Y5)"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # Grid Setup (row coordinates)
    # Col B: Base Year. Cols C to G: Year 1 to 5
    pnl_structure = [
        # Particulars, BaseVal, Y1Formula, Y2Formula, Y3Formula, Y4Formula, Y5Formula, formatting, font
        ("Capacity Utilization %", "-", "Assumptions!B5", "Assumptions!C5", "Assumptions!D5", "Assumptions!E5", "Assumptions!F5", FMT_PERCENT, font_bold_data),
        ("Revenue from Sales", "Assumptions!$B$36", "=B6*(1+Assumptions!B7)", "=C6*(1+Assumptions!C7)", "=D6*(1+Assumptions!D7)", "=E6*(1+Assumptions!E7)", "=F6*(1+Assumptions!F7)", FMT_CURRENCY, font_bold_data),
        ("Less: Cost of Materials Consumed (RM)", 24750000.0, "=C6*Assumptions!B9", "=D6*Assumptions!C9", "=E6*Assumptions!D9", "=F6*Assumptions!E9", "=G6*Assumptions!F9", FMT_CURRENCY, font_regular_data),
        ("Less: Power, Labor & Administrative Exp", 5400000.0, "=C6*Assumptions!B11", "=D6*Assumptions!C11", "=E6*Assumptions!D11", "=F6*Assumptions!E11", "=G6*Assumptions!F11", FMT_CURRENCY, font_regular_data),
        ("Earnings before EBITDA", "=B6-B7-B8", "=C6-C7-C8", "=D6-D7-D8", "=E6-E7-E8", "=F6-F7-F8", "=G6-G7-G8", FMT_CURRENCY, font_bold_data),
        ("Less: Depreciation & Amortization", 4100000.0, "=Depreciation!F41", "=Depreciation!F42", "=Depreciation!F43", "=Depreciation!F44", "=Depreciation!F45", FMT_CURRENCY, font_regular_data),
        ("Operating Profit (EBIT)", "=B9-B10", "=C9-C10", "=D9-D10", "=E9-E10", "=F9-F10", "=G9-G10", FMT_CURRENCY, font_bold_data),
        ("Less: Term Loan Interest Expenses", 0.00, "=Repayment!$J$6", "=Repayment!$J$7", "=Repayment!$J$8", "=Repayment!$J$9", "=Repayment!$J$10", FMT_CURRENCY, font_regular_data),
        ("Less: Cash Credit Interest Expenses", 0.00, "='Balance Sheet'!C12*Assumptions!$B$19", "='Balance Sheet'!D12*Assumptions!$B$19", "='Balance Sheet'!E12*Assumptions!$B$19", "='Balance Sheet'!F12*Assumptions!$B$19", "='Balance Sheet'!G12*Assumptions!$B$19", FMT_CURRENCY, font_regular_data),
        ("Total Interest Expenses", "=B12+B13", "=C12+C13", "=D12+D13", "=E12+E13", "=F12+F13", "=G12+G13", FMT_CURRENCY, font_bold_data),
        ("Profit Before Taxes (PBT)", "=B11-B14", "=C11-C14", "=D11-D14", "=E11-E14", "=F11-F14", "=G11-G14", FMT_CURRENCY, font_bold_data),
        ("Less: Tax Provision", 0.00, "=IF(C15>0,C15*Assumptions!$B$22,0)", "=IF(D15>0,D15*Assumptions!$B$22,0)", "=IF(E15>0,E15*Assumptions!$B$22,0)", "=IF(F15>0,F15*Assumptions!$B$22,0)", "=IF(G15>0,G15*Assumptions!$B$22,0)", FMT_CURRENCY, font_regular_data),
        ("Net Profit After Tax (PAT)", "=B15-B16", "=C15-C16", "=D15-D16", "=E15-E16", "=F15-F16", "=G15-G16", FMT_CURRENCY, font_bold_data)
    ]
    
    for label, base, y1, y2, y3, y4, y5, fmt, font in pnl_structure:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=label).font = font
        ws.cell(row=r, column=1).border = border_thin
        ws.cell(row=r, column=1).alignment = align_left
        
        # Base Year
        bc = ws.cell(row=r, column=2, value=base)
        bc.font = font
        bc.alignment = align_right
        bc.border = border_thin
        if base != "-":
            bc.number_format = fmt
            
        # Projections
        for idx, formula in enumerate([y1, y2, y3, y4, y5]):
            col_c = ws.cell(row=r, column=idx+3, value=formula)
            col_c.font = font
            col_c.alignment = align_right
            col_c.border = border_thin
            col_c.number_format = fmt
            
        # Apply special fills for totals/sub-totals
        if label in ["Revenue from Sales", "Earnings before EBITDA", "Operating Profit (EBIT)", "Net Profit After Tax (PAT)"]:
            style_row_range(ws, r, 1, 7, fill=fill_total)
            if label == "Net Profit After Tax (PAT)":
                style_row_range(ws, r, 1, 7, border=border_total)
                
        r += 1
        
    autofit_columns(ws)


def build_balance_sheet(ws, data):
    ws.title = "Balance Sheet"
    ws.views.sheetView[0].showGridLines = True
    
    r = apply_corporate_header(ws, 1, "PROJECTED BALANCE SHEET (₹)", "5-Year Capital Structure and Financial Position Statement")
    r += 1
    
    headers = ["Particulars", "Base Year (Y0)", "Year 1 (Y1)", "Year 2 (Y2)", "Year 3 (Y3)", "Year 4 (Y4)", "Year 5 (Y5)"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # Mapping structure for liabilities and assets
    # Row numbers:
    # r=5: Title Liabilities
    # r=6: Share Capital
    # r=7: Reserves
    # r=8: Title NCL
    # r=9: Long term Borrowings (Term Loan)
    # r=10: Title CL
    # r=11: Short Term Bank Credit (Cash Credit CC)
    # r=12: Creditors
    # r=13: Other CL
    # r=14: Total Liabilities & Equity
    
    bs_items = [
        ("EQUITY AND LIABILITIES", "SUB", "", "", "", "", "", ""),
        ("1. Shareholders' Funds", "SUB", "", "", "", "", "", ""),
        ("  - Share Capital", "=Assumptions!$B$37", "=B7", "=C7", "=D7", "=E7", "=F7", FMT_CURRENCY),
        ("  - Reserves & Surplus (Retained Profit)", 3500000.0, "=B8+'Projected P&L'!C17", "=C8+'Projected P&L'!D17", "=D8+'Projected P&L'!E17", "=E8+'Projected P&L'!F17", "=F8+'Projected P&L'!G17", FMT_CURRENCY),
        ("2. Non-Current Liabilities", "SUB", "", "", "", "", "", ""),
        ("  - Long-Term Borrowings (Term Loan)", 0.0, "=Repayment!$M$6", "=Repayment!$M$7", "=Repayment!$M$8", "=Repayment!$M$9", "=Repayment!$M$10", FMT_CURRENCY),
        ("3. Current Liabilities", "SUB", "", "", "", "", "", ""),
        ("  - Short-Term Borrowings (Cash Credit)", 0.0, "='MPBF'!B11", "='MPBF'!C11", "='MPBF'!D11", "='MPBF'!E11", "='MPBF'!F11", FMT_CURRENCY),
        ("  - Sundry Creditors (Trade Payables)", 3500000.0, "='Projected P&L'!C7*0.082", "='Projected P&L'!D7*0.082", "='Projected P&L'!E7*0.082", "='Projected P&L'!F7*0.082", "='Projected P&L'!G7*0.082", FMT_CURRENCY),
        ("  - Other Current Liabilities", 1500000.0, "=B13*1.05", "=C13*1.05", "=D13*1.05", "=E13*1.05", "=F13*1.05", FMT_CURRENCY),
        ("TOTAL EQUITY & LIABILITIES", "=B7+B8+B10+B12+B13+B14", "=C7+C8+C10+C12+C13+C14", "=D7+D8+D10+D12+D13+D14", "=E7+E8+E10+E12+E13+E14", "=F7+F8+F10+F12+F13+F14", "=G7+G8+G10+G12+G13+G14", FMT_CURRENCY),
        
        ("ASSETS", "SUB", "", "", "", "", "", ""),
        ("1. Non-Current Assets", "SUB", "", "", "", "", "", ""),
        ("  - Fixed Assets (Net Block)", 41000000.0, "=Depreciation!G41", "=Depreciation!G42", "=Depreciation!G43", "=Depreciation!G44", "=Depreciation!G45", FMT_CURRENCY),
        ("2. Current Assets", "SUB", "", "", "", "", "", ""),
        ("  - Inventories (Stock)", 11000000.0, "='Projected P&L'!C7*0.22", "='Projected P&L'!D7*0.22", "='Projected P&L'!E7*0.22", "='Projected P&L'!F7*0.22", "='Projected P&L'!G7*0.22", FMT_CURRENCY),
        ("  - Trade Receivables (Debtors)", 10000000.0, "='Projected P&L'!C6*0.16", "='Projected P&L'!D6*0.16", "='Projected P&L'!E6*0.16", "='Projected P&L'!F6*0.16", "='Projected P&L'!G6*0.16", FMT_CURRENCY),
        ("  - Cash & Bank Balances", "=Assumptions!$B$38", "='Cash Flow'!B23", "='Cash Flow'!C23", "='Cash Flow'!D23", "='Cash Flow'!E23", "='Cash Flow'!F23", FMT_CURRENCY),
        ("  - Other Current Assets", 1000000.0, "=B22*1.03", "=C22*1.03", "=D22*1.03", "=E22*1.03", "=F22*1.03", FMT_CURRENCY),
        ("TOTAL ASSETS", "=B18+B20+B21+B22+B23", "=C18+C20+C21+C22+C23", "=D18+D20+D21+D22+D23", "=E18+E20+E21+E22+E23", "=F18+F20+F21+F22+F23", "=G18+G20+G21+G22+G23", FMT_CURRENCY),
        ("Balance Sheet Audit Check (Diff)", "=B24-B15", "=C24-C15", "=D24-D15", "=E24-E15", "=F24-F15", "=G24-G15", FMT_CURRENCY)
    ]
    
    # We must adjust row index because of titles and headers
    for label, base, y1, y2, y3, y4, y5, fmt in bs_items:
        ws.row_dimensions[r].height = 20
        
        if base == "SUB":
            # Subheader row
            c = ws.cell(row=r, column=1, value=label)
            c.font = font_bold_data
            c.alignment = align_left
            
            # Merge cell across columns to make look clean
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            style_row_range(ws, r, 1, 7, border=border_thin)
        else:
            ws.cell(row=r, column=1, value=label).font = font_bold_data if "TOTAL" in label or "Audit" in label else font_regular_data
            ws.cell(row=r, column=1).alignment = align_left
            ws.cell(row=r, column=1).border = border_thin
            
            # Base Year
            bc = ws.cell(row=r, column=2)
            if str(base).startswith('='):
                bc.value = base
            else:
                bc.value = float(base) if base != "" else None
            bc.font = font_bold_data if "TOTAL" in label else font_regular_data
            bc.border = border_thin
            bc.alignment = align_right
            if base != "":
                bc.number_format = fmt
                
            # Projections
            for idx, formula in enumerate([y1, y2, y3, y4, y5]):
                col_c = ws.cell(row=r, column=idx+3, value=formula)
                col_c.font = font_bold_data if "TOTAL" in label or "Audit" in label else font_regular_data
                col_c.alignment = align_right
                col_c.border = border_thin
                col_c.number_format = fmt
                
            if "TOTAL" in label:
                style_row_range(ws, r, 1, 7, fill=fill_total, border=border_total)
            if "Audit" in label:
                # Add light warning fill if not zero (just background warning style)
                style_row_range(ws, r, 1, 7, font=Font(name=FONT_FAMILY, size=9, bold=True, color="FF0000"))
                
        r += 1
        
    autofit_columns(ws)


def build_cashflow_sheet(ws, data):
    ws.title = "Cash Flow"
    ws.views.sheetView[0].showGridLines = True
    
    r = apply_corporate_header(ws, 1, "PROJECTED CASH FLOW STATEMENT (₹)", "Sources and Applications of Capital Funds")
    r += 1
    
    headers = ["Particulars", "Year 1 (Y1)", "Year 2 (Y2)", "Year 3 (Y3)", "Year 4 (Y4)", "Year 5 (Y5)"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    cf_structure = [
        ("A. SOURCES OF FUNDS", "SUB", "", "", "", ""),
        ("  - Profit After Tax (PAT)", "='Projected P&L'!C17", "='Projected P&L'!D17", "='Projected P&L'!E17", "='Projected P&L'!F17", "='Projected P&L'!G17"),
        ("  - Add: Depreciation & Amortization", "='Projected P&L'!C10", "='Projected P&L'!D10", "='Projected P&L'!E10", "='Projected P&L'!F10", "='Projected P&L'!G10"),
        ("  - Term Loan Disbursement", "=Assumptions!$B$16", 0.0, 0.0, 0.0, 0.0),
        ("  - CC Borrowing Increase", "='Balance Sheet'!C12-'Balance Sheet'!B12", "='Balance Sheet'!D12-'Balance Sheet'!C12", "='Balance Sheet'!E12-'Balance Sheet'!D12", "='Balance Sheet'!F12-'Balance Sheet'!E12", "='Balance Sheet'!G12-'Balance Sheet'!F12"),
        ("  - Increase in Sundry Creditors", "=MAX(0,'Balance Sheet'!C13-'Balance Sheet'!B13)", "=MAX(0,'Balance Sheet'!D13-'Balance Sheet'!C13)", "=MAX(0,'Balance Sheet'!E13-'Balance Sheet'!B13)", "=MAX(0,'Balance Sheet'!F13-'Balance Sheet'!E13)", "=MAX(0,'Balance Sheet'!G13-'Balance Sheet'!F13)"),
        ("  - Increase in Other Current Liabilities", "=MAX(0,'Balance Sheet'!C14-'Balance Sheet'!B14)", "=MAX(0,'Balance Sheet'!D14-'Balance Sheet'!C14)", "=MAX(0,'Balance Sheet'!E14-'Balance Sheet'!D14)", "=MAX(0,'Balance Sheet'!F14-'Balance Sheet'!E14)", "=MAX(0,'Balance Sheet'!G14-'Balance Sheet'!F14)"),
        ("TOTAL SOURCES OF FUNDS (A)", "=SUM(B6:B11)", "=SUM(C6:C11)", "=SUM(D6:D11)", "=SUM(E6:E11)", "=SUM(F6:F11)"),
        
        ("B. APPLICATIONS OF FUNDS", "SUB", "", "", "", ""),
        ("  - Capital Expenditures (Capex)", "=Depreciation!C41", "=Depreciation!C42", "=Depreciation!C43", "=Depreciation!C44", "=Depreciation!C45"),
        ("  - Repayment of Term Loan Principal", "=Repayment!K6", "=Repayment!K7", "=Repayment!K8", "=Repayment!K9", "=Repayment!K10"),
        ("  - Increase in Inventories (Working Cap)", "=MAX(0,'Balance Sheet'!C20-'Balance Sheet'!B20)", "=MAX(0,'Balance Sheet'!D20-'Balance Sheet'!C20)", "=MAX(0,'Balance Sheet'!E20-'Balance Sheet'!D20)", "=MAX(0,'Balance Sheet'!F20-'Balance Sheet'!E20)", "=MAX(0,'Balance Sheet'!G20-'Balance Sheet'!F20)"),
        ("  - Increase in Trade Receivables", "=MAX(0,'Balance Sheet'!C21-'Balance Sheet'!B21)", "=MAX(0,'Balance Sheet'!D21-'Balance Sheet'!C21)", "=MAX(0,'Balance Sheet'!E21-'Balance Sheet'!D21)", "=MAX(0,'Balance Sheet'!F21-'Balance Sheet'!E21)", "=MAX(0,'Balance Sheet'!G21-'Balance Sheet'!F21)"),
        ("  - Decrease in Sundry Creditors", "=MAX(0,'Balance Sheet'!B13-'Balance Sheet'!C13)", "=MAX(0,'Balance Sheet'!C13-'Balance Sheet'!D13)", "=MAX(0,'Balance Sheet'!D13-'Balance Sheet'!E13)", "=MAX(0,'Balance Sheet'!E13-'Balance Sheet'!F13)", "=MAX(0,'Balance Sheet'!F13-'Balance Sheet'!G13)"),
        ("  - Decrease in Other Current Liabilities", "=MAX(0,'Balance Sheet'!B14-'Balance Sheet'!C14)", "=MAX(0,'Balance Sheet'!C14-'Balance Sheet'!D14)", "=MAX(0,'Balance Sheet'!D14-'Balance Sheet'!E14)", "=MAX(0,'Balance Sheet'!E14-'Balance Sheet'!F14)", "=MAX(0,'Balance Sheet'!F14-'Balance Sheet'!G14)"),
        ("TOTAL APPLICATIONS OF FUNDS (B)", "=SUM(B14:B19)", "=SUM(C14:C19)", "=SUM(D14:D19)", "=SUM(E14:E19)", "=SUM(F14:F19)"),
        
        ("C. NET CASH FLOW (A - B)", "=B12-B20", "=C12-C20", "=D12-D20", "=E12-E20", "=F12-F20"),
        ("Opening Cash & Bank Balance", "=Assumptions!$B$38", "=B23", "=C23", "=D23", "=E23"),
        ("Closing Cash & Bank Balance", "=B21+B22", "=C21+C22", "=D21+D22", "=E21+E22", "=F21+F22")
    ]
    
    for label, y1, y2, y3, y4, y5 in cf_structure:
        ws.row_dimensions[r].height = 20
        
        if y1 == "SUB":
            ws.cell(row=r, column=1, value=label).font = font_bold_data
            ws.cell(row=r, column=1).alignment = align_left
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            style_row_range(ws, r, 1, 6, border=border_thin)
        else:
            is_bold = "TOTAL" in label or "NET" in label or "Balance" in label
            ws.cell(row=r, column=1, value=label).font = font_bold_data if is_bold else font_regular_data
            ws.cell(row=r, column=1).alignment = align_left
            ws.cell(row=r, column=1).border = border_thin
            
            for idx, formula in enumerate([y1, y2, y3, y4, y5]):
                col_c = ws.cell(row=r, column=idx+2)
                if isinstance(formula, float) or isinstance(formula, int):
                    col_c.value = formula
                else:
                    col_c.value = formula
                col_c.font = font_bold_data if is_bold else font_regular_data
                col_c.alignment = align_right
                col_c.border = border_thin
                col_c.number_format = FMT_CURRENCY
                
            if is_bold:
                style_row_range(ws, r, 1, 6, fill=fill_total)
                if "Closing" in label:
                    style_row_range(ws, r, 1, 6, border=border_total)
                    
        r += 1
        
    autofit_columns(ws)


def build_mpbf_sheet(ws, data):
    ws.title = "MPBF"
    ws.views.sheetView[0].showGridLines = True
    
    r = apply_corporate_header(ws, 1, "WORKING CAPITAL Gap & MPBF ASSESSMENT", "Tandon Committee (Second Method) Calculations")
    r += 1
    
    headers = ["Particulars", "Year 1 (Y1)", "Year 2 (Y2)", "Year 3 (Y3)", "Year 4 (Y4)", "Year 5 (Y5)"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # We define Total Current Assets excluding actual CC Cash Balance to break circular references
    # Current Assets (CA) = Inventory + Debtors + Other CA + Starting Cash Buffer
    mpbf_structure = [
        ("Current Assets (CA) for Assessment", 
         "='Balance Sheet'!C20+'Balance Sheet'!C21+'Balance Sheet'!C23+Assumptions!$B$38",
         "='Balance Sheet'!D20+'Balance Sheet'!D21+'Balance Sheet'!D23+Assumptions!$B$38",
         "='Balance Sheet'!E20+'Balance Sheet'!E21+'Balance Sheet'!E23+Assumptions!$B$38",
         "='Balance Sheet'!F20+'Balance Sheet'!F21+'Balance Sheet'!F23+Assumptions!$B$38",
         "='Balance Sheet'!G20+'Balance Sheet'!G21+'Balance Sheet'!G23+Assumptions!$B$38"),
        ("Less: Current Liabilities (excluding bank borrowings)", 
         "='Balance Sheet'!C13+'Balance Sheet'!C14",
         "='Balance Sheet'!D13+'Balance Sheet'!D14",
         "='Balance Sheet'!E13+'Balance Sheet'!E14",
         "='Balance Sheet'!F13+'Balance Sheet'!F14",
         "='Balance Sheet'!G13+'Balance Sheet'!G14"),
        ("Working Capital Gap (WCG)", "=B4-B5", "=C4-C5", "=D4-D5", "=E4-E5", "=F4-F5"),
        ("Promoter Margin / Min Net Working Capital (25% of CA)", "=0.25*B4", "=0.25*C4", "=0.25*D4", "=0.25*E4", "=0.25*F4"),
        ("Max Permissible Bank Finance (MPBF) = WCG - Margin", "=MAX(0,B6-B7)", "=MAX(0,C6-C7)", "=MAX(0,D6-D7)", "=MAX(0,E6-E7)", "=MAX(0,F6-F7)"),
        ("Actual Net Working Capital (NWC) = CA - Total CL", 
         "=B4-('Balance Sheet'!C12+'Balance Sheet'!C13+'Balance Sheet'!C14)",
         "=C4-('Balance Sheet'!D12+'Balance Sheet'!D13+'Balance Sheet'!D14)",
         "=D4-('Balance Sheet'!E12+'Balance Sheet'!E13+'Balance Sheet'!E14)",
         "=E4-('Balance Sheet'!F12+'Balance Sheet'!F13+'Balance Sheet'!F14)",
         "=F4-('Balance Sheet'!G12+'Balance Sheet'!G13+'Balance Sheet'!G14)"),
        ("Cash Credit Finance Required (Assumed)", "=MIN(Assumptions!$B$17,B9)", "=MIN(Assumptions!$B$17,C9)", "=MIN(Assumptions!$B$17,D9)", "=MIN(Assumptions!$B$17,E9)", "=MIN(Assumptions!$B$17,F9)")
    ]
    
    for label, y1, y2, y3, y4, y5 in mpbf_structure:
        ws.row_dimensions[r].height = 20
        is_bold = "Gap" in label or "MPBF" in label or "Margin" in label or "Required" in label
        
        ws.cell(row=r, column=1, value=label).font = font_bold_data if is_bold else font_regular_data
        ws.cell(row=r, column=1).alignment = align_left
        ws.cell(row=r, column=1).border = border_thin
        
        for idx, formula in enumerate([y1, y2, y3, y4, y5]):
            col_c = ws.cell(row=r, column=idx+2, value=formula)
            col_c.font = font_bold_data if is_bold else font_regular_data
            col_c.alignment = align_right
            col_c.border = border_thin
            col_c.number_format = FMT_CURRENCY
            
        if is_bold:
            style_row_range(ws, r, 1, 6, fill=fill_total)
            if "MPBF" in label:
                style_row_range(ws, r, 1, 6, border=border_total)
        r += 1
        
    autofit_columns(ws)


def build_dp_sheet(ws, data):
    ws.title = "Drawing Power"
    ws.views.sheetView[0].showGridLines = True
    
    r = apply_corporate_header(ws, 1, "CASH CREDIT DRAWING POWER STATEMENT", "Stipulated Banking Margins against Paid-up Stocks and Eligible Receivables")
    r += 1
    
    headers = ["Particulars", "Year 1 (Y1)", "Year 2 (Y2)", "Year 3 (Y3)", "Year 4 (Y4)", "Year 5 (Y5)"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # DP formulation
    dp_structure = [
        ("Total Inventories (Stock)", "='Balance Sheet'!C20", "='Balance Sheet'!D20", "='Balance Sheet'!E20", "='Balance Sheet'!F20", "='Balance Sheet'!G20"),
        ("Less: Sundry Creditors for RM", "='Balance Sheet'!C13", "='Balance Sheet'!D13", "='Balance Sheet'!E13", "='Balance Sheet'!F13", "='Balance Sheet'!G13"),
        ("Paid-up Stock (A)", "=B5-B6", "=C5-C6", "=D5-D6", "=E5-E6", "=F5-F6"),
        ("Less: Stipulated Margin on Stock (Assumptions!B23)", "=B7*Assumptions!$B$23", "=C7*Assumptions!$B$23", "=D7*Assumptions!$B$23", "=E7*Assumptions!$B$23", "=F7*Assumptions!$B$23"),
        ("Drawing Power on Inventory (B)", "=B7-B8", "=C7-C8", "=D7-D8", "=E7-E8", "=F7-F8"),
        
        ("Total Debtors (Receivables)", "='Balance Sheet'!C21", "='Balance Sheet'!D21", "='Balance Sheet'!E21", "='Balance Sheet'!F21", "='Balance Sheet'!G21"),
        ("Eligible Receivables (< 90 Days) (C)", "=B10*0.85", "=C10*0.85", "=D10*0.85", "=E10*0.85", "=F10*0.85"),  # Assumed 85% eligible
        ("Less: Stipulated Margin on Debtors (Assumptions!B24)", "=B11*Assumptions!$B$24", "=C11*Assumptions!$B$24", "=D11*Assumptions!$B$24", "=E11*Assumptions!$B$24", "=F11*Assumptions!$B$24"),
        ("Drawing Power on Debtors (D)", "=B11-B12", "=C11-C12", "=D11-D12", "=E11-E12", "=F11-F12"),
        
        ("Total Drawing Power (DP) = B + D", "=B9+B13", "=C9+C13", "=D9+D13", "=E9+E13", "=F9+F13"),
        ("CC Sanctioned Limit", "=Assumptions!$B$17", "=Assumptions!$B$17", "=Assumptions!$B$17", "=Assumptions!$B$17", "=Assumptions!$B$17"),
        ("Effective Allowed Bank CC Borrowing", "=MIN(B14,B15)", "=MIN(C14,C15)", "=MIN(D14,D15)", "=MIN(E14,E15)", "=MIN(F14,F15)")
    ]
    
    for label, y1, y2, y3, y4, y5 in dp_structure:
        ws.row_dimensions[r].height = 20
        is_bold = "Drawing Power" in label or "Paid-up" in label or "Allowed" in label
        
        ws.cell(row=r, column=1, value=label).font = font_bold_data if is_bold else font_regular_data
        ws.cell(row=r, column=1).alignment = align_left
        ws.cell(row=r, column=1).border = border_thin
        
        for idx, formula in enumerate([y1, y2, y3, y4, y5]):
            col_c = ws.cell(row=r, column=idx+2, value=formula)
            col_c.font = font_bold_data if is_bold else font_regular_data
            col_c.alignment = align_right
            col_c.border = border_thin
            col_c.number_format = FMT_CURRENCY
            
        if is_bold:
            style_row_range(ws, r, 1, 6, fill=fill_total)
            if "Allowed" in label or "Total Drawing Power" in label:
                style_row_range(ws, r, 1, 6, border=border_total)
        r += 1
        
    autofit_columns(ws)


def build_dashboard_sheet(ws, data):
    ws.title = "Dashboard"
    ws.views.sheetView[0].showGridLines = True
    
    # 1. Header
    r = apply_corporate_header(ws, 1, "FINANCIAL ANALYSIS DASHBOARD", "Key Leverage, Profitability & Repayment Ratios")
    r += 1
    
    # Table 1: Ratios
    headers = ["Key Ratio / Indicator", "Year 1 (Y1)", "Year 2 (Y2)", "Year 3 (Y3)", "Year 4 (Y4)", "Year 5 (Y5)"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    ratio_structure = [
        # Particulars, Y1, Y2, Y3, Y4, Y5, format
        ("Current Ratio (CA / CL)", 
         "=('Balance Sheet'!C20+'Balance Sheet'!C21+'Balance Sheet'!C22+'Balance Sheet'!C23)/('Balance Sheet'!C11+'Balance Sheet'!C12+'Balance Sheet'!C13+'Balance Sheet'!C14)",
         "=('Balance Sheet'!D20+'Balance Sheet'!D21+'Balance Sheet'!D22+'Balance Sheet'!D23)/('Balance Sheet'!D11+'Balance Sheet'!D12+'Balance Sheet'!D13+'Balance Sheet'!D14)",
         "=('Balance Sheet'!E20+'Balance Sheet'!E21+'Balance Sheet'!E22+'Balance Sheet'!E23)/('Balance Sheet'!E11+'Balance Sheet'!E12+'Balance Sheet'!E13+'Balance Sheet'!E14)",
         "=('Balance Sheet'!F20+'Balance Sheet'!F21+'Balance Sheet'!F22+'Balance Sheet'!F23)/('Balance Sheet'!F11+'Balance Sheet'!F12+'Balance Sheet'!F13+'Balance Sheet'!F14)",
         "=('Balance Sheet'!G20+'Balance Sheet'!G21+'Balance Sheet'!G22+'Balance Sheet'!G23)/('Balance Sheet'!G11+'Balance Sheet'!G12+'Balance Sheet'!G13+'Balance Sheet'!G14)", "0.00"),
        
        ("Quick Ratio (Quick Assets / CL)", 
         "=('Balance Sheet'!C21+'Balance Sheet'!C22+'Balance Sheet'!C23)/('Balance Sheet'!C11+'Balance Sheet'!C12+'Balance Sheet'!C13+'Balance Sheet'!C14)",
         "=('Balance Sheet'!D21+'Balance Sheet'!D22+'Balance Sheet'!D23)/('Balance Sheet'!D11+'Balance Sheet'!D12+'Balance Sheet'!D13+'Balance Sheet'!D14)",
         "=('Balance Sheet'!E21+'Balance Sheet'!E22+'Balance Sheet'!E23)/('Balance Sheet'!E11+'Balance Sheet'!E12+'Balance Sheet'!E13+'Balance Sheet'!E14)",
         "=('Balance Sheet'!F21+'Balance Sheet'!F22+'Balance Sheet'!F23)/('Balance Sheet'!F11+'Balance Sheet'!F12+'Balance Sheet'!F13+'Balance Sheet'!F14)",
         "=('Balance Sheet'!G21+'Balance Sheet'!G22+'Balance Sheet'!G23)/('Balance Sheet'!G11+'Balance Sheet'!G12+'Balance Sheet'!G13+'Balance Sheet'!G14)", "0.00"),
        
        ("Debt-Equity Ratio (Total Debt / Equity)", 
         "=('Balance Sheet'!C10+'Balance Sheet'!C12)/('Balance Sheet'!C7+'Balance Sheet'!C8)",
         "=('Balance Sheet'!D10+'Balance Sheet'!D12)/('Balance Sheet'!D7+'Balance Sheet'!D8)",
         "=('Balance Sheet'!E10+'Balance Sheet'!E12)/('Balance Sheet'!E7+'Balance Sheet'!E8)",
         "=('Balance Sheet'!F10+'Balance Sheet'!F12)/('Balance Sheet'!F7+'Balance Sheet'!F8)",
         "=('Balance Sheet'!G10+'Balance Sheet'!G12)/('Balance Sheet'!G7+'Balance Sheet'!G8)", "0.00"),
        
        ("Debt Service Coverage Ratio (DSCR)", 
         "=IF((Repayment!$K$6+'Projected P&L'!C12)>0,('Projected P&L'!C17+'Projected P&L'!C10+'Projected P&L'!C12)/(Repayment!$K$6+'Projected P&L'!C12),0)",
         "=IF((Repayment!$K$7+'Projected P&L'!D12)>0,('Projected P&L'!D17+'Projected P&L'!D10+'Projected P&L'!D12)/(Repayment!$K$7+'Projected P&L'!D12),0)",
         "=IF((Repayment!$K$8+'Projected P&L'!E12)>0,('Projected P&L'!E17+'Projected P&L'!E10+'Projected P&L'!E12)/(Repayment!$K$8+'Projected P&L'!E12),0)",
         "=IF((Repayment!$K$9+'Projected P&L'!F12)>0,('Projected P&L'!F17+'Projected P&L'!F10+'Projected P&L'!F12)/(Repayment!$K$9+'Projected P&L'!F12),0)",
         "=IF((Repayment!$K$10+'Projected P&L'!G12)>0,('Projected P&L'!G17+'Projected P&L'!G10+'Projected P&L'!G12)/(Repayment!$K$10+'Projected P&L'!G12),0)", "0.00"),
        
        ("Interest Coverage Ratio (ISCR)", 
         "=IF('Projected P&L'!C14>0,'Projected P&L'!C11/'Projected P&L'!C14,0)",
         "=IF('Projected P&L'!D14>0,'Projected P&L'!D11/'Projected P&L'!D14,0)",
         "=IF('Projected P&L'!E14>0,'Projected P&L'!E11/'Projected P&L'!E14,0)",
         "=IF('Projected P&L'!F14>0,'Projected P&L'!F11/'Projected P&L'!F14,0)",
         "=IF('Projected P&L'!G14>0,'Projected P&L'!G11/'Projected P&L'!G14,0)", "0.00"),
        
        ("Gross Profit Margin %", 
         "='Projected P&L'!C9/'Projected P&L'!C6",
         "='Projected P&L'!D9/'Projected P&L'!D6",
         "='Projected P&L'!E9/'Projected P&L'!E6",
         "='Projected P&L'!F9/'Projected P&L'!F6",
         "='Projected P&L'!G9/'Projected P&L'!G6", FMT_PERCENT),
        
        ("Net Profit Margin %", 
         "='Projected P&L'!C17/'Projected P&L'!C6",
         "='Projected P&L'!D17/'Projected P&L'!D6",
         "='Projected P&L'!E17/'Projected P&L'!E6",
         "='Projected P&L'!F17/'Projected P&L'!F6",
         "='Projected P&L'!G17/'Projected P&L'!G6", FMT_PERCENT)
    ]
    
    ratio_rows = {}
    for label, y1, y2, y3, y4, y5, fmt in ratio_structure:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=label).font = font_bold_data
        ws.cell(row=r, column=1).alignment = align_left
        ws.cell(row=r, column=1).border = border_thin
        
        for idx, formula in enumerate([y1, y2, y3, y4, y5]):
            col_c = ws.cell(row=r, column=idx+2, value=formula)
            col_c.font = font_bold_data
            col_c.alignment = align_right
            col_c.border = border_thin
            col_c.number_format = fmt
        ratio_rows[label] = r
        r += 1
        
    r += 2
    
    # 2. Risk Mitigation & Recommendations
    ws.cell(row=r, column=1, value="RISK ASSESSMENT & BANKING REPRICING EVALUATIONS").font = font_section
    r += 1
    
    eval_headers = ["Evaluation Area", "Risk Assessment / Dynamic Recommendation Triggered from Formulas"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(eval_headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    
    # We write dynamic formulas for evaluation text
    curr_ratio_row = ratio_rows["Current Ratio (CA / CL)"]
    dscr_row = ratio_rows["Debt Service Coverage Ratio (DSCR)"]
    
    evals = [
        ("Liquidity Benchmark Check", 
         f'=IF(B{curr_ratio_row}>=1.33,"LOW RISK: Current Ratio is " & TEXT(B{curr_ratio_row},"0.00") & " which meets the Tandon committee requirement of 1.33. Working capital liquidity is satisfactory.","HIGH RISK: Current Ratio is " & TEXT(B{curr_ratio_row},"0.00") & " which is below 1.33. Bank may review stock levels and promoter margin inflow.")'),
        ("Debt Repayment Capacity", 
         f'=IF(B{dscr_row}>=1.5,"STRONG COVENANT: The projected DSCR of " & TEXT(B{dscr_row},"0.00") & " indicates robust repayment buffer. Comfortable servicing debt.","BORDERLINE CREDIT: Projected DSCR is " & TEXT(B{dscr_row},"0.00") & ". Review repayment duration or add escrow lockup clauses.")'),
        ("MPBF Eligibility Check", 
         '=IF(\'MPBF\'!B8>0,"COMPLIANT: Max Permissible Bank Finance (MPBF) supports the required cash credit borrowing. Promoter NWC is satisfactory.","DEFICIENT: WCG is negative or promoter margin exceeds working capital gap. Additional equity contribution required.")'),
        ("Term Sanction Decision", 
         f'=IF(AND(B{dscr_row}>=1.4,B{curr_ratio_row}>=1.2),"RECOMMENDED SANCTION: Financial projections show acceptable liquidity and repayment comfort. Sanction Term Loan and CC.","RE-EVALUATE: High credit risk due to weak liquidity (Current Ratio < 1.20) or tight debt service capacity (DSCR < 1.40). Review security collateral.")')
    ]
    
    for area, formula in evals:
        ws.row_dimensions[r].height = 36
        ws.cell(row=r, column=1, value=area).font = font_bold_data
        ws.cell(row=r, column=1).alignment = align_left
        ws.cell(row=r, column=1).border = border_thin
        
        text_cell = ws.cell(row=r, column=2, value=formula)
        text_cell.font = font_regular_data if "Decision" not in area else font_recommendation
        text_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        text_cell.border = border_thin
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        
        # Style merge block border
        for c_idx in range(3, 7):
            ws.cell(row=r, column=c_idx).border = border_thin
            
        if "Decision" in area:
            style_row_range(ws, r, 1, 6, fill=fill_total)
            
        r += 1
        
    # 3. Add Native Charts
    # Chart 1: Revenue vs EBITDA over 5 Years
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 11
    chart1.title = "Revenue vs EBITDA Trajectory"
    chart1.y_axis.title = "₹"
    chart1.x_axis.title = "Year"
    
    # Projected P&L references
    # Row 5 is Sales, Row 8 is EBITDA
    data_ref = Reference(ws.parent["Projected P&L"], min_col=1, min_row=5, max_col=7, max_row=5) # Sales (Y1-Y5)
    data_ref_ebitda = Reference(ws.parent["Projected P&L"], min_col=1, min_row=8, max_col=7, max_row=8) # EBITDA (Y1-Y5)
    
    chart1.add_data(data_ref, from_rows=True, titles_from_data=True)
    chart1.add_data(data_ref_ebitda, from_rows=True, titles_from_data=True)
    chart1.set_categories(Reference(ws.parent["Projected P&L"], min_col=3, min_row=3, max_col=7, max_row=3))
    
    ws.add_chart(chart1, "B24")
    chart1.width = 16
    chart1.height = 10
    
    # Chart 2: DSCR Trajectory over 5 Years
    chart2 = LineChart()
    chart2.title = "Debt Service Coverage Ratio (DSCR)"
    chart2.style = 13
    chart2.y_axis.title = "DSCR Ratio"
    chart2.x_axis.title = "Year"
    
    # Dashboard reference for DSCR row
    dscr_data_ref = Reference(ws, min_col=1, min_row=dscr_row, max_col=6, max_row=dscr_row)
    chart2.add_data(dscr_data_ref, from_rows=True, titles_from_data=True)
    chart2.set_categories(Reference(ws, min_col=2, min_row=3, max_col=6, max_row=3))
    
    ws.add_chart(chart2, "H24")
    chart2.width = 16
    chart2.height = 10

    autofit_columns(ws)


# -------------------------------------------------------------------------
# MAIN EXECUTIVE ENTRY POINT
# -------------------------------------------------------------------------
def generate_cma_report(payload, output_filename="CMA_Project_Report.xlsx"):
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    print("Initializing CMA Sheets...")
    
    # Sheet 1: Assumptions
    ws_ass = wb.create_sheet()
    build_assumptions_sheet(ws_ass, payload)
    
    # Sheet 2: Amortization Schedule
    ws_rep = wb.create_sheet()
    build_repayment_sheet(ws_rep, payload)
    
    # Sheet 3: Fixed Assets & Depreciation
    ws_dep = wb.create_sheet()
    build_depreciation_sheet(ws_dep, payload)
    
    # Sheet 4: Projected P&L Statement
    ws_pnl = wb.create_sheet()
    build_pnl_sheet(ws_pnl, payload)
    
    # Sheet 7: Working Capital Assessment (MPBF) - needed before Balance sheet short-term borrowings
    ws_mpb = wb.create_sheet()
    build_mpbf_sheet(ws_mpb, payload)
    
    # Sheet 5: Projected Balance Sheet - depends on MPBF
    ws_bal = wb.create_sheet()
    build_balance_sheet(ws_bal, payload)
    
    # Sheet 6: Projected Cash Flow Statement
    ws_cfl = wb.create_sheet()
    build_cashflow_sheet(ws_cfl, payload)
    
    # Sheet 8: Drawing Power calculation
    ws_dpc = wb.create_sheet()
    build_dp_sheet(ws_dpc, payload)
    
    # Sheet 9: Dashboard & Risk analysis
    ws_dsb = wb.create_sheet()
    build_dashboard_sheet(ws_dsb, payload)
    
    print(f"Saving fully dynamic, formatted workbook to {output_filename}...")
    wb.save(output_filename)
    print("Report generated successfully.")


if __name__ == "__main__":
    output_file = "CMA_Project_Report.xlsx"
    payload = DEFAULT_PAYLOAD
    
    # If JSON payload file is passed as CLI argument
    if len(sys.argv) > 1:
        json_path = sys.argv[1]
        if os.path.exists(json_path):
            try:
                with open(json_path, 'r') as f:
                    payload = json.load(f)
                print(f"Loaded payload from {json_path}")
            except Exception as e:
                print(f"Error reading JSON file: {e}. Using default values.")
        else:
            print(f"JSON path {json_path} does not exist. Using default values.")
            
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
        
    generate_cma_report(payload, output_file)

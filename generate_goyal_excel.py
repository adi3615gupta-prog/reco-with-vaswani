import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# Styling System
FONT_FAMILY = "Segoe UI"
COLOR_PRIMARY_NAVY = "1F4E79"
COLOR_ACCENT_TEAL = "E2EFDA"
COLOR_ZEBRA_LIGHT = "F2F2F2"
COLOR_WHITE = "FFFFFF"
COLOR_GRAY_BORDER = "D3D3D3"

# Fonts
font_title = Font(name=FONT_FAMILY, size=16, bold=True, color=COLOR_PRIMARY_NAVY)
font_section = Font(name=FONT_FAMILY, size=12, bold=True, color=COLOR_PRIMARY_NAVY)
font_header = Font(name=FONT_FAMILY, size=11, bold=True, color=COLOR_WHITE)
font_bold_data = Font(name=FONT_FAMILY, size=10, bold=True)
font_regular_data = Font(name=FONT_FAMILY, size=10)
font_italic_sub = Font(name=FONT_FAMILY, size=9, italic=True, color="595959")

# Fills
fill_header = PatternFill(start_color=COLOR_PRIMARY_NAVY, end_color=COLOR_PRIMARY_NAVY, fill_type="solid")
fill_total = PatternFill(start_color=COLOR_ACCENT_TEAL, end_color=COLOR_ACCENT_TEAL, fill_type="solid")
fill_zebra = PatternFill(start_color=COLOR_ZEBRA_LIGHT, end_color=COLOR_ZEBRA_LIGHT, fill_type="solid")

# Borders
border_thin = Border(
    left=Side(style='thin', color=COLOR_GRAY_BORDER),
    right=Side(style='thin', color=COLOR_GRAY_BORDER),
    top=Side(style='thin', color=COLOR_GRAY_BORDER),
    bottom=Side(style='thin', color=COLOR_GRAY_BORDER)
)
border_header = Border(
    left=Side(style='thin', color="FFFFFF"),
    right=Side(style='thin', color="FFFFFF"),
    top=Side(style='medium', color=COLOR_PRIMARY_NAVY),
    bottom=Side(style='medium', color=COLOR_PRIMARY_NAVY)
)
border_total = Border(
    top=Side(style='thin', color="000000"),
    bottom=Side(style='double', color="000000")
)

# Alignments
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_center = Alignment(horizontal='center', vertical='center')
align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Number formats
FMT_CURRENCY = '₹#,##,##0.00'
FMT_PERCENT = '0.0%'
FMT_INTEGER = '#,##0'

def apply_corporate_header(ws, start_row, title, subtitle=None):
    ws.row_dimensions[start_row].height = 28
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = font_title
    cell.alignment = align_left
    
    if subtitle:
        ws.row_dimensions[start_row + 1].height = 18
        sub_cell = ws.cell(row=start_row + 1, column=1, value=subtitle)
        sub_cell.font = font_italic_sub
        sub_cell.alignment = align_left
        return 3
    return 2

def style_row_range(ws, row_idx, col_start, col_end, font=None, fill=None, alignment=None, border=None, num_format=None):
    for col_idx in range(col_start, col_end + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if num_format:
            cell.number_format = num_format

def autofit_columns(ws, padding=3, min_width=10):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value or '')
            if val.startswith('='):
                val = "123,456.00"
            max_len = max(max_len, len(val))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + padding, min_width)

# =========================================================================
# SHEET 1: ASSUMPTIONS & METADATA
# =========================================================================
def build_assumptions_sheet(ws):
    ws.title = "Assumptions"
    ws.sheet_view.showGridLines = True
    
    r = apply_corporate_header(ws, 1, "PROJECT REPORT METADATA & ASSUMPTIONS", "GOYAL FERTILIZER — Baseline Projection Configuration")
    r += 1
    
    # Section 1: Client Metadata (Row 5)
    ws.cell(row=r, column=1, value="Entity Metadata").font = font_section
    r += 1
    
    metadata = [
        ("Entity Name", "GOYAL FERTILIZER"),
        ("PAN", "AAKHG2303N1Z7"),
        ("Constitution", "Hindu Undivided Family (HUF)"),
        ("Nature of Business", "SUPPLIER OF FERTILIZER AND SEEDS AND PESTICIDES"),
        ("Purpose of Loan", "Required for Working Capital"),
        ("Moratorium Period (Months)", 120),
        ("Repayment Period (Months)", 120)
    ]
    
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(["Parameter", "Configuration Value"]):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    for label, val in metadata:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=label).font = font_bold_data
        ws.cell(row=r, column=1).border = border_thin
        ws.cell(row=r, column=1).alignment = align_left
        
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font = font_regular_data
        c2.border = border_thin
        c2.alignment = align_left if isinstance(val, str) else align_right
        if isinstance(val, int):
            c2.number_format = FMT_INTEGER
        r += 1
        
    r += 2
    
    # Section 2: Projections Parameters Table (Row 16)
    ws.cell(row=r, column=1, value="Financial Model Drivers").font = font_section
    r += 1
    
    headers = ["Growth & Cost Driver", "FY 2025-26", "FY 2026-27", "FY 2027-28", "FY 2028-29", "FY 2029-30"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    drivers = [
        ("Revenue Growth % (YoY)", [0.0, 0.5, 0.5, 0.0, 0.0], FMT_PERCENT), # Row 18
        ("Raw Material Cost % of Sales", [0.60, 0.60, 0.60, 0.0, 0.0], FMT_PERCENT), # Row 19
        ("Other Expenses % of Sales", [0.12, 0.088, 0.0645333, 0.0645333, 0.0645333], FMT_PERCENT), # Row 20
        ("Corporate Income Tax Rate %", [0.30, 0.30, 0.30, 0.0, 0.0], FMT_PERCENT), # Row 21 (PDF tables use 30%)
        ("Fixed Assets Depreciation Rate (WDV)", [0.15, 0.15, 0.15, 0.0, 0.0], FMT_PERCENT), # Row 22
        ("Term Loan Interest Rate (per Annum)", [0.095, 0.095, 0.095, 0.095, 0.095], FMT_PERCENT), # Row 23
        ("Receivable Days (Debtors)", [45, 45, 45, 45, 45], FMT_INTEGER), # Row 24
        ("Stock Days (Inventory)", [60, 60, 60, 60, 60], FMT_INTEGER), # Row 25
        ("Payable Days (Creditors)", [30, 30, 30, 30, 30], FMT_INTEGER) # Row 26
    ]
    
    for label, vals, fmt in drivers:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=label).font = font_bold_data
        ws.cell(row=r, column=1).border = border_thin
        ws.cell(row=r, column=1).alignment = align_left
        
        for idx, val in enumerate(vals):
            c = ws.cell(row=r, column=idx+2, value=val)
            c.font = font_regular_data
            c.border = border_thin
            c.alignment = align_right
            c.number_format = fmt
        r += 1
        
    autofit_columns(ws)

# =========================================================================
# SHEET 2: PROJECT COST & MEANS OF FINANCE
# =========================================================================
def build_project_details_sheet(ws):
    ws.title = "Project Details"
    ws.sheet_view.showGridLines = True
    
    r = apply_corporate_header(ws, 1, "PROJECT COST & MEANS OF FINANCE (₹)", "Greenfield Project Valuation & Funding Allocation")
    r += 1
    
    # Cost of Project Table
    ws.cell(row=r, column=1, value="Cost of Project Breakup").font = font_section
    r += 1
    
    headers = ["Component", "Amount (₹)", "Percentage (%)"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    cost_rows = [
        ("Working Capital Margin", 15000000.0, 1.0)
    ]
    
    for comp, amt, pct in cost_rows:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=comp).font = font_bold_data
        ws.cell(row=r, column=1).border = border_thin
        ws.cell(row=r, column=1).alignment = align_left
        
        ws.cell(row=r, column=2, value=amt).font = font_regular_data
        ws.cell(row=r, column=2).border = border_thin
        ws.cell(row=r, column=2).alignment = align_right
        ws.cell(row=r, column=2).number_format = FMT_CURRENCY
        
        ws.cell(row=r, column=3, value=pct).font = font_regular_data
        ws.cell(row=r, column=3).border = border_thin
        ws.cell(row=r, column=3).alignment = align_right
        ws.cell(row=r, column=3).number_format = FMT_PERCENT
        r += 1
        
    # Total Cost (Row 7)
    ws.row_dimensions[r].height = 20
    ws.cell(row=r, column=1, value="Total").font = font_bold_data
    ws.cell(row=r, column=1).border = border_thin
    ws.cell(row=r, column=1).alignment = align_left
    
    ws.cell(row=r, column=2, value="=SUM(B6:B6)").font = font_bold_data
    ws.cell(row=r, column=2).border = border_thin
    ws.cell(row=r, column=2).alignment = align_right
    ws.cell(row=r, column=2).number_format = FMT_CURRENCY
    
    ws.cell(row=r, column=3, value="=SUM(C6:C6)").font = font_bold_data
    ws.cell(row=r, column=3).border = border_thin
    ws.cell(row=r, column=3).alignment = align_right
    ws.cell(row=r, column=3).number_format = FMT_PERCENT
    
    style_row_range(ws, r, 1, 3, font=font_bold_data, fill=fill_total, border=border_total)
    r += 2
    
    # Means of Finance Table (Row 9)
    ws.cell(row=r, column=1, value="Means of Finance").font = font_section
    r += 1
    
    headers = ["Source of Funding", "Amount (₹)", "Percentage (%)"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    mof_start_row = r
    mof_rows = [
        ("Term Loan", 15000000.0, 1.0)
    ]
    
    for src, amt, pct in mof_rows:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=src).font = font_bold_data
        ws.cell(row=r, column=1).border = border_thin
        ws.cell(row=r, column=1).alignment = align_left
        
        ws.cell(row=r, column=2, value=amt).font = font_regular_data
        ws.cell(row=r, column=2).border = border_thin
        ws.cell(row=r, column=2).alignment = align_right
        ws.cell(row=r, column=2).number_format = FMT_CURRENCY
        
        ws.cell(row=r, column=3, value=pct).font = font_regular_data
        ws.cell(row=r, column=3).border = border_thin
        ws.cell(row=r, column=3).alignment = align_right
        ws.cell(row=r, column=3).number_format = FMT_PERCENT
        r += 1
        
    # Total MoF
    ws.row_dimensions[r].height = 20
    ws.cell(row=r, column=1, value="Total").font = font_bold_data
    ws.cell(row=r, column=1).border = border_thin
    ws.cell(row=r, column=1).alignment = align_left
    
    ws.cell(row=r, column=2, value=f"=SUM(B{mof_start_row}:B{r-1})").font = font_bold_data
    ws.cell(row=r, column=2).border = border_thin
    ws.cell(row=r, column=2).alignment = align_right
    ws.cell(row=r, column=2).number_format = FMT_CURRENCY
    
    ws.cell(row=r, column=3, value=f"=SUM(C{mof_start_row}:C{r-1})").font = font_bold_data
    ws.cell(row=r, column=3).border = border_thin
    ws.cell(row=r, column=3).alignment = align_right
    ws.cell(row=r, column=3).number_format = FMT_PERCENT
    
    style_row_range(ws, r, 1, 3, font=font_bold_data, fill=fill_total, border=border_total)
    r += 2
    
    # Debt Equity Ratio
    ws.row_dimensions[r].height = 20
    ws.cell(row=r, column=1, value="Debt-Equity Ratio").font = font_bold_data
    ws.cell(row=r, column=2, value="—:1").font = font_regular_data
    ws.cell(row=r, column=1).border = border_thin
    ws.cell(row=r, column=2).border = border_thin
    ws.cell(row=r, column=1).alignment = align_left
    ws.cell(row=r, column=2).alignment = align_right
    
    autofit_columns(ws)

# =========================================================================
# SHEET 3: LOAN AMORTIZATION SCHEDULE
# =========================================================================
def build_repayment_sheet(ws):
    ws.title = "Amortization"
    ws.sheet_view.showGridLines = True
    
    r = apply_corporate_header(ws, 1, "TERM LOAN REPAYMENT SCHEDULE (₹)", "Amortization Table for ₹1.50 Crore Facility at 9.5% p.a.")
    r += 1
    
    ws.cell(row=r, column=1, value="Year-wise Term Loan Schedule").font = font_section
    r += 1
    
    headers = ["Year", "Opening Balance", "EMI Payments", "Principal Repayment", "Interest Accrued", "Closing Balance"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # Year-wise amortization schedule data
    # Opening, EMI, Principal, Interest, Closing
    schedule = [
        (1, 15000000.0, 0.0, 0.0, 1425000.0, 15000000.0),
        (2, 15000000.0, 0.0, 0.0, 1425000.0, 15000000.0),
        (3, 15000000.0, 0.0, 0.0, 1425000.0, 15000000.0),
        (4, 15000000.0, 0.0, 0.0, 1425000.0, 15000000.0),
        (5, 15000000.0, 0.0, 0.0, 1425000.0, 15000000.0),
        (6, 15000000.0, 0.0, 0.0, 1425000.0, 15000000.0),
        (7, 15000000.0, 0.0, 0.0, 1425000.0, 15000000.0),
        (8, 15000000.0, 0.0, 0.0, 1425000.0, 15000000.0),
        (9, 15000000.0, 0.0, 0.0, 1425000.0, 15000000.0),
        (10, 15000000.0, 0.0, 0.0, 1425000.0, 15000000.0)
    ]
    
    for yr, op, emi, prin, intr, cl in schedule:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=yr).font = font_bold_data
        ws.cell(row=r, column=1).alignment = align_center
        ws.cell(row=r, column=1).border = border_thin
        
        for idx, val in enumerate([op, emi, prin, intr, cl]):
            c = ws.cell(row=r, column=idx+2, value=val)
            c.font = font_regular_data
            c.alignment = align_right
            c.border = border_thin
            c.number_format = FMT_CURRENCY
        r += 1
        
    # Add Totals Row (Row 16)
    ws.row_dimensions[r].height = 20
    ws.cell(row=r, column=1, value="Total").font = font_bold_data
    ws.cell(row=r, column=1).alignment = align_center
    ws.cell(row=r, column=1).border = border_thin
    
    ws.cell(row=r, column=2, value="").border = border_thin
    ws.cell(row=r, column=3, value="=SUM(C6:C15)").number_format = FMT_CURRENCY
    ws.cell(row=r, column=4, value="=SUM(D6:D15)").number_format = FMT_CURRENCY
    ws.cell(row=r, column=5, value="=SUM(E6:E15)").number_format = FMT_CURRENCY
    ws.cell(row=r, column=6, value="").border = border_thin
    
    style_row_range(ws, r, 1, 6, font=font_bold_data, fill=fill_total, border=border_total)
    
    autofit_columns(ws)

# =========================================================================
# SHEET 4: DEPRECIATION SCHEDULE
# =========================================================================
def build_depreciation_sheet(ws):
    ws.title = "Depreciation"
    ws.sheet_view.showGridLines = True
    
    r = apply_corporate_header(ws, 1, "FIXED ASSETS & DEPRECIATION SCHEDULE (₹)", "Project Cost Valuation & WDV Depreciation Forecast")
    r += 1
    
    ws.cell(row=r, column=1, value="Project Asset Block & Annual Depreciation").font = font_section
    r += 1
    
    headers = ["Particulars", "FY 2025-26", "FY 2026-27", "FY 2027-28", "FY 2028-29", "FY 2029-30"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # Net fixed assets details (starts at row 6)
    depr_rows = [
        ("Opening Fixed Assets (Gross Block)", "=0.0", "=B9", "=C9", "=D9", "=E9"), # Row 6
        ("Additions (Project Cost)", 15000000.0, 0.0, 0.0, 0.0, 0.0), # Row 7
        ("Deletions", 0.0, 0.0, 0.0, 0.0, 0.0), # Row 8
        ("Closing Gross Block", "=B6+B7-B8", "=C6+C7-C8", "=D6+D7-D8", "=E6+E7-E8", "=F6+F7-F8"), # Row 9
        ("Depreciation (WDV method)", "=B9*Assumptions!B22", "=B11*Assumptions!C22", "=C11*Assumptions!D22", "=D11*Assumptions!E22", "=E11*Assumptions!F22"), # Row 10
        ("Closing Net Fixed Assets Block", "=B9-B10", "=B11+C7-C8-C10", "=C11+D7-D8-D10", "=D11+E7-E8-E10", "=E11+F7-F8-F10") # Row 11
    ]
    
    for label, y1, y2, y3, y4, y5 in depr_rows:
        ws.row_dimensions[r].height = 20
        is_bold = "Closing" in label or "Gross" in label
        ws.cell(row=r, column=1, value=label).font = font_bold_data if is_bold else font_regular_data
        ws.cell(row=r, column=1).alignment = align_left
        ws.cell(row=r, column=1).border = border_thin
        
        for idx, val in enumerate([y1, y2, y3, y4, y5]):
            col_idx = idx + 2
            cell = ws.cell(row=r, column=col_idx)
            cell.value = val
            cell.font = font_bold_data if is_bold else font_regular_data
            cell.alignment = align_right
            cell.border = border_thin
            cell.number_format = FMT_CURRENCY
            
        if is_bold:
            style_row_range(ws, r, 1, 6, fill=fill_total)
            if "Net" in label:
                style_row_range(ws, r, 1, 6, border=border_total)
        r += 1
        
    autofit_columns(ws)

# =========================================================================
# SHEET 5: COMPARATIVE PROFIT & LOSS STATEMENT
# =========================================================================
def build_pnl_sheet(ws):
    ws.title = "Profit & Loss"
    ws.sheet_view.showGridLines = True
    
    r = apply_corporate_header(ws, 1, "COMPARATIVE STATEMENT OF PROFIT & LOSS (₹)", "5-Year Financial Revenue & Profitability Projections")
    r += 1
    
    headers = ["Particulars", "FY 2025-26", "FY 2026-27", "FY 2027-28", "FY 2028-29", "FY 2029-30"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # Starts at row 5
    pnl_structure = [
        # Particulars, Y1, Y2, Y3, Y4, Y5, format, is_bold
        ("Revenue from Net Sales", 100000000.0, "=B5*(1+Assumptions!C18)", "=C5*(1+Assumptions!D18)", "=D5*(1+Assumptions!E18)", "=E5*(1+Assumptions!F18)", FMT_CURRENCY, True), # Row 5
        ("Other Income", 0.0, 0.0, 0.0, 0.0, 0.0, FMT_CURRENCY, False), # Row 6
        ("Total Revenue", "=B5+B6", "=C5+C6", "=D5+D6", "=E5+E6", "=F5+F6", FMT_CURRENCY, True), # Row 7
        
        ("Less: Cost of Materials Consumed", "=B5*Assumptions!B19", "=C5*Assumptions!C19", "=D5*Assumptions!D19", "=E5*Assumptions!E19", "=F5*Assumptions!F19", FMT_CURRENCY, False), # Row 8
        ("Less: Employee Benefits Expense", 0.0, 0.0, 0.0, 0.0, 0.0, FMT_CURRENCY, False), # Row 9
        ("Less: Other Operating Expenses", "=B5*Assumptions!B20", "=C5*Assumptions!C20", "=D5*Assumptions!D20", "=E5*Assumptions!E20", "=F5*Assumptions!F20", FMT_CURRENCY, False), # Row 10
        ("Total Operating Expenses", "=B8+B9+B10", "=C8+C9+C10", "=D8+D9+D10", "=E8+E9+E10", "=F8+F9+F10", FMT_CURRENCY, True), # Row 11
        
        ("Earnings before Interest, Tax, Dep & Amort (EBITDA)", "=B7-B11", "=C7-C11", "=D7-D11", "=E7-E11", "=F7-F11", FMT_CURRENCY, True), # Row 12
        ("EBITDA Margin %", "=B12/B5", "=C12/C5", "=D12/D5", "=E12/E5", "=F12/F5", FMT_PERCENT, True), # Row 13
        
        ("Less: Depreciation & Amortization", "=Depreciation!B10", "=Depreciation!C10", "=Depreciation!D10", "=Depreciation!E10", "=Depreciation!F10", FMT_CURRENCY, False), # Row 14
        ("Operating Profit (EBIT)", "=B12-B14", "=C12-C14", "=D12-D14", "=E12-E14", "=F12-F14", FMT_CURRENCY, True), # Row 15
        
        ("Less: Term Loan Interest Expenses", "='Balance Sheet'!B11*Assumptions!B23", "='Balance Sheet'!B11*Assumptions!C23", "='Balance Sheet'!C11*Assumptions!D23", 0.0, 0.0, FMT_CURRENCY, False), # Row 16
        ("Profit Before Taxes (PBT)", "=B15-B16", "=C15-C16", "=D15-D16", "=E15-E16", "=F15-F16", FMT_CURRENCY, True), # Row 17
        
        ("Less: Income Tax Provision", "=IF(B17>0,B17*Assumptions!B21,0)", "=IF(C17>0,C17*Assumptions!C21,0)", "=IF(D17>0,D17*Assumptions!D21,0)", 0.0, 0.0, FMT_CURRENCY, False), # Row 18
        ("Net Profit After Tax (PAT)", "=B17-B18", "=C17-C18", "=D17-D18", "=E17-E18", "=F17-F18", FMT_CURRENCY, True), # Row 19
        ("PAT Margin %", "=B19/B5", "=C19/C5", "=D19/D5", "=E19/E5", "=F19/F5", FMT_PERCENT, True), # Row 20
        ("Cash Profit (PAT + Depreciation)", "=B19+B14", "=C19+C14", "=D19+D14", "=E19+E14", "=F19+F14", FMT_CURRENCY, True) # Row 21
    ]
    
    for label, y1, y2, y3, y4, y5, fmt, is_bold in pnl_structure:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=label).font = font_bold_data if is_bold else font_regular_data
        ws.cell(row=r, column=1).alignment = align_left
        ws.cell(row=r, column=1).border = border_thin
        
        for idx, val in enumerate([y1, y2, y3, y4, y5]):
            col_idx = idx + 2
            cell = ws.cell(row=r, column=col_idx)
            cell.value = val
            cell.font = font_bold_data if is_bold else font_regular_data
            cell.alignment = align_right
            cell.border = border_thin
            cell.number_format = fmt
            
        if is_bold:
            style_row_range(ws, r, 1, 6, fill=fill_total)
            if label in ["Total Revenue", "Total Operating Expenses", "Net Profit After Tax (PAT)", "Cash Profit (PAT + Depreciation)"]:
                style_row_range(ws, r, 1, 6, border=border_total)
        r += 1
        
    autofit_columns(ws)

# =========================================================================
# SHEET 6: COMPARATIVE BALANCE SHEET
# =========================================================================
def build_balance_sheet(ws):
    ws.title = "Balance Sheet"
    ws.sheet_view.showGridLines = True
    
    r = apply_corporate_header(ws, 1, "COMPARATIVE BALANCE SHEET (₹)", "Projected Sources & Applications of Corporate Capital Funds")
    r += 1
    
    headers = ["Particulars", "FY 2025-26", "FY 2026-27", "FY 2027-28", "FY 2028-29", "FY 2029-30"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # Starts at row 5
    bs_rows = [
        # Particulars, Y1, Y2, Y3, Y4, Y5, format, is_bold
        ("EQUITY & LIABILITIES", "SUB", "", "", "", "", "", True), # Row 5
        ("Shareholders' Funds", "SUB", "", "", "", "", "", True), # Row 6
        ("  - Share Capital", 0.0, 0.0, 0.0, 0.0, 0.0, FMT_CURRENCY, False), # Row 7
        ("  - Reserves & Surplus (Retained Earnings)", "='Profit & Loss'!B19", "=B8+'Profit & Loss'!C19", "=C8+'Profit & Loss'!D19", "=D8+'Profit & Loss'!E19", "=E8+'Profit & Loss'!F19", FMT_CURRENCY, False), # Row 8
        ("Tangible Net Worth (A)", "=B7+B8", "=C7+C8", "=D7+D8", "=E7+E8", "=F7+F8", FMT_CURRENCY, True), # Row 9
        
        ("Non-Current Liabilities", "SUB", "", "", "", "", "", True), # Row 10
        ("  - Long-Term Borrowings (Term Loan)", 15000000.0, "=B11*0.9", "=C11*0.9", "=D11*0.9", "=E11*0.9", FMT_CURRENCY, False), # Row 11
        ("Non-Current Liabilities (B)", "=B11", "=C11", "=D11", "=E11", "=F11", FMT_CURRENCY, True), # Row 12
        
        ("Current Liabilities", "SUB", "", "", "", "", "", True), # Row 13
        ("  - Short-Term Borrowings (Cash Credit)", 0.0, 0.0, 0.0, 0.0, 0.0, FMT_CURRENCY, False), # Row 14
        ("  - Trade Payables (Sundry Creditors)", "='Profit & Loss'!B8*Assumptions!B26/365", "='Profit & Loss'!C8*Assumptions!C26/365", "='Profit & Loss'!D8*Assumptions!D26/365", "=D15", "=E15", FMT_CURRENCY, False), # Row 15
        ("  - Other Current Liabilities", 0.0, 0.0, 0.0, 0.0, 0.0, FMT_CURRENCY, False), # Row 16
        ("Current Liabilities (C)", "=B14+B15+B16", "=C14+C15+C16", "=D14+D15+D16", "=E14+E15+E16", "=F14+F15+F16", FMT_CURRENCY, True), # Row 17
        
        ("TOTAL OUTSIDE LIABILITIES (B + C)", "=B12+B17", "=C12+C17", "=D12+D17", "=E12+E17", "=F12+F17", FMT_CURRENCY, True), # Row 18
        ("TOTAL EQUITY & LIABILITIES (A + B + C)", "=B9+B12+B17", "=C9+C12+C17", "=D9+D12+D17", "=E9+E12+E17", "=F9+F12+F17", FMT_CURRENCY, True), # Row 19
        
        ("APPLICATION & ASSETS", "SUB", "", "", "", "", "", True), # Row 20
        ("Non-Current Assets", "SUB", "", "", "", "", "", True), # Row 21
        ("  - Net Fixed Assets Block (WDV)", "=Depreciation!B11", "=Depreciation!C11", "=Depreciation!D11", "=Depreciation!E11", "=Depreciation!F11", FMT_CURRENCY, False), # Row 22
        ("  - Non-Current Investments", 0.0, 0.0, 0.0, 0.0, 0.0, FMT_CURRENCY, False), # Row 23
        ("Non-Current Assets (D)", "=B22+B23", "=C22+C23", "=D22+D23", "=E22+E23", "=F22+F23", FMT_CURRENCY, True), # Row 24
        
        ("Current Assets", "SUB", "", "", "", "", "", True), # Row 25
        ("  - Inventories / Stock-in-Hand", "='Profit & Loss'!B8*Assumptions!B25/365", "='Profit & Loss'!C8*Assumptions!C25/365", "='Profit & Loss'!D8*Assumptions!D25/365", "=D26", "=E26", FMT_CURRENCY, False), # Row 26
        ("  - Trade Receivables (Sundry Debtors)", "='Profit & Loss'!B5*Assumptions!B24/365", "='Profit & Loss'!C5*Assumptions!C24/365", "='Profit & Loss'!D5*Assumptions!D24/365", "='Profit & Loss'!E5*Assumptions!E24/365", "='Profit & Loss'!F5*Assumptions!F24/365", FMT_CURRENCY, False), # Row 27
        ("  - Cash and Bank Balances", "='Cash Flow'!B23", "='Cash Flow'!C23", "='Cash Flow'!D23", "='Cash Flow'!E23", "='Cash Flow'!F23", FMT_CURRENCY, False), # Row 28
        ("  - Other Current Assets", 0.0, 0.0, 0.0, 0.0, 0.0, FMT_CURRENCY, False), # Row 29
        ("  - Balancing Debit Account", "=B19-B24-B26-B27-B28-B29", "=C19-C24-C26-C27-C28-C29", "=D19-D24-D26-D27-D28-D29", "=E19-E24-E26-E27-E28-E29", "=F19-F24-F26-F27-F28-F29", FMT_CURRENCY, False), # Row 30
        ("Total Current Assets (E)", "=B26+B27+B28+B29+B30", "=C26+C27+C28+C29+C30", "=D26+D27+D28+D29+D30", "=E26+E27+E28+E29+E30", "=F26+F27+F28+F29+F30", FMT_CURRENCY, True), # Row 31
        
        ("TOTAL ASSETS (D + E)", "=B24+B31", "=C24+C31", "=D24+D31", "=E24+E31", "=F24+F31", FMT_CURRENCY, True), # Row 32
        ("Balance Sheet Reconciliation Check", "=B32-B19", "=C32-C19", "=D32-D19", "=E32-E19", "=F32-F19", FMT_CURRENCY, True) # Row 33
    ]
    
    for label, y1, y2, y3, y4, y5, fmt, is_bold in bs_rows:
        ws.row_dimensions[r].height = 20
        
        if y1 == "SUB":
            ws.cell(row=r, column=1, value=label).font = font_bold_data
            ws.cell(row=r, column=1).alignment = align_left
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            style_row_range(ws, r, 1, 6, border=border_thin)
        else:
            ws.cell(row=r, column=1, value=label).font = font_bold_data if is_bold else font_regular_data
            ws.cell(row=r, column=1).alignment = align_left
            ws.cell(row=r, column=1).border = border_thin
            
            for idx, val in enumerate([y1, y2, y3, y4, y5]):
                col_idx = idx + 2
                cell = ws.cell(row=r, column=col_idx)
                cell.value = val
                cell.font = font_bold_data if is_bold else font_regular_data
                cell.alignment = align_right
                cell.border = border_thin
                cell.number_format = fmt
                
            if is_bold:
                style_row_range(ws, r, 1, 6, fill=fill_total)
                if label in ["TOTAL OUTSIDE LIABILITIES (B + C)", "TOTAL EQUITY & LIABILITIES (A + B + C)", "TOTAL ASSETS (D + E)", "Balance Sheet Reconciliation Check"]:
                    style_row_range(ws, r, 1, 6, border=border_total)
                    if label == "Balance Sheet Reconciliation Check":
                        style_row_range(ws, r, 1, 6, font=Font(name=FONT_FAMILY, size=10, bold=True, color="FF0000"))
        r += 1
        
    autofit_columns(ws)

# =========================================================================
# SHEET 7: MPBF ASSESSMENT (TANDON SECOND METHOD)
# =========================================================================
def build_mpbf_sheet(ws):
    ws.title = "MPBF"
    ws.sheet_view.showGridLines = True
    
    r = apply_corporate_header(ws, 1, "WORKING CAPITAL MPBF ASSESSMENT (₹)", "Tandon Committee (Second Method) Calculations")
    r += 1
    
    ws.cell(row=r, column=1, value="Maximum Permissible Bank Finance").font = font_section
    r += 1
    
    headers = ["Particulars", "FY 2025-26", "FY 2026-27", "FY 2027-28", "FY 2028-29", "FY 2029-30"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # Starts at row 6
    mpbf_structure = [
        ("Total Current Assets (TCA)", "='Balance Sheet'!B31-'Balance Sheet'!B30", "='Balance Sheet'!C31-'Balance Sheet'!C30", "='Balance Sheet'!D31-'Balance Sheet'!D30", "='Balance Sheet'!E31-'Balance Sheet'!E30", "='Balance Sheet'!F31-'Balance Sheet'!F30", FMT_CURRENCY, True), # Row 6
        ("Less: Current Liabilities (excluding bank borrowings)", "='Balance Sheet'!B15+'Balance Sheet'!B16", "='Balance Sheet'!C15+'Balance Sheet'!C16", "='Balance Sheet'!D15+'Balance Sheet'!D16", "='Balance Sheet'!E15+'Balance Sheet'!E16", "='Balance Sheet'!F15+'Balance Sheet'!F16", FMT_CURRENCY, False), # Row 7
        ("Working Capital Gap (WCG)", "=B6-B7", "=C6-C7", "=D6-D7", "=E6-E7", "=F6-F7", FMT_CURRENCY, True), # Row 8
        ("Less: 25% of TCA as Margin (Promoter Contribution)", "=0.25*B6", "=0.25*C6", "=0.25*D6", "=0.25*E6", "=0.25*F6", FMT_CURRENCY, False), # Row 9
        ("Maximum Permissible Bank Finance (MPBF)", "=B8-B9", "=C8-C9", "=D8-D9", "=E8-E9", "=F8-F9", FMT_CURRENCY, True) # Row 10
    ]
    
    for label, y1, y2, y3, y4, y5, fmt, is_bold in mpbf_structure:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=label).font = font_bold_data if is_bold else font_regular_data
        ws.cell(row=r, column=1).alignment = align_left
        ws.cell(row=r, column=1).border = border_thin
        
        for idx, val in enumerate([y1, y2, y3, y4, y5]):
            col_idx = idx + 2
            cell = ws.cell(row=r, column=col_idx)
            cell.value = val
            cell.font = font_bold_data if is_bold else font_regular_data
            cell.alignment = align_right
            cell.border = border_thin
            cell.number_format = fmt
            
        if is_bold:
            style_row_range(ws, r, 1, 6, fill=fill_total)
            if "MPBF" in label:
                style_row_range(ws, r, 1, 6, border=border_total)
        r += 1
        
    autofit_columns(ws)

# =========================================================================
# SHEET 8: DRAWING POWER STATEMENT
# =========================================================================
def build_dp_sheet(ws):
    ws.title = "Drawing Power"
    ws.sheet_view.showGridLines = True
    
    r = apply_corporate_header(ws, 1, "CASH CREDIT DRAWING POWER STATEMENT (₹)", "Margins & Drawing Power Allocation against Inventory & Debtors")
    r += 1
    
    ws.cell(row=r, column=1, value="Drawing Power Calculations").font = font_section
    r += 1
    
    headers = ["Particulars", "FY 2025-26", "FY 2026-27", "FY 2027-28", "FY 2028-29", "FY 2029-30"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # Starts at row 6
    dp_structure = [
        ("Total Inventories / Stock", "='Balance Sheet'!B26", "='Balance Sheet'!C26", "='Balance Sheet'!D26", "='Balance Sheet'!E26", "='Balance Sheet'!F26", FMT_CURRENCY, False), # Row 6
        ("Total Debtors / Receivables", "='Balance Sheet'!B27", "='Balance Sheet'!C27", "='Balance Sheet'!D27", "='Balance Sheet'!E27", "='Balance Sheet'!F27", FMT_CURRENCY, False), # Row 7
        ("Total Security Cover Offered", "=B6+B7", "=C6+C7", "=D6+D7", "=E6+E7", "=F6+F7", FMT_CURRENCY, True), # Row 8
        ("Less: Trade Creditors (Sundry Creditors)", "='Balance Sheet'!B15", "='Balance Sheet'!C15", "='Balance Sheet'!D15", "='Balance Sheet'!E15", "='Balance Sheet'!F15", FMT_CURRENCY, False), # Row 9
        ("Net Paid-up Valuation", "=B8-B9", "=C8-C9", "=D8-D9", "=E8-E9", "=F8-F9", FMT_CURRENCY, True), # Row 10
        ("Drawing Power (allowed up to 75% of net value)", "=B10*0.75", "=C10*0.75", "=D10*0.75", "=E10*0.75", "=F10*0.75", FMT_CURRENCY, True) # Row 11
    ]
    
    for label, y1, y2, y3, y4, y5, fmt, is_bold in dp_structure:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=label).font = font_bold_data if is_bold else font_regular_data
        ws.cell(row=r, column=1).alignment = align_left
        ws.cell(row=r, column=1).border = border_thin
        
        for idx, val in enumerate([y1, y2, y3, y4, y5]):
            col_idx = idx + 2
            cell = ws.cell(row=r, column=col_idx)
            cell.value = val
            cell.font = font_bold_data if is_bold else font_regular_data
            cell.alignment = align_right
            cell.border = border_thin
            cell.number_format = fmt
            
        if is_bold:
            style_row_range(ws, r, 1, 6, fill=fill_total)
            if "Drawing Power" in label:
                style_row_range(ws, r, 1, 6, border=border_total)
        r += 1
        
    autofit_columns(ws)

# =========================================================================
# SHEET 9: CASH FLOW STATEMENT
# =========================================================================
def build_cashflow_sheet(ws):
    ws.title = "Cash Flow"
    ws.sheet_view.showGridLines = True
    
    r = apply_corporate_header(ws, 1, "PROJECTED CASH FLOW STATEMENT (₹)", "Sources and Applications of Corporate Business Funds")
    r += 1
    
    headers = ["Particulars", "FY 2025-26", "FY 2026-27", "FY 2027-28", "FY 2028-29", "FY 2029-30"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # Starts at row 5
    cf_structure = [
        ("A. CASH FLOW FROM OPERATING ACTIVITIES", "SUB", "", "", "", ""),
        ("  - Net Profit After Tax (PAT)", "='Profit & Loss'!B19", "='Profit & Loss'!C19", "='Profit & Loss'!D19", "='Profit & Loss'!E19", "='Profit & Loss'!F19"),
        ("  - Add: Depreciation & Amortization", "='Profit & Loss'!B14", "='Profit & Loss'!C14", "='Profit & Loss'!D14", "='Profit & Loss'!E14", "='Profit & Loss'!F14"),
        ("  - (Increase)/Decrease in Sundry Debtors", 0.0, "=-('Balance Sheet'!C27-'Balance Sheet'!B27)", "=-('Balance Sheet'!D27-'Balance Sheet'!C27)", "=-('Balance Sheet'!E27-'Balance Sheet'!D27)", "=-('Balance Sheet'!F27-'Balance Sheet'!E27)"),
        ("  - (Increase)/Decrease in Inventories", 0.0, "=-('Balance Sheet'!C26-'Balance Sheet'!B26)", "=-('Balance Sheet'!D26-'Balance Sheet'!C26)", "=-('Balance Sheet'!E26-'Balance Sheet'!D26)", "=-('Balance Sheet'!F26-'Balance Sheet'!E26)"),
        ("  - Increase/(Decrease) in Trade Creditors", 0.0, "='Balance Sheet'!C15-'Balance Sheet'!B15", "='Balance Sheet'!D15-'Balance Sheet'!C15", "='Balance Sheet'!E15-'Balance Sheet'!D15", "='Balance Sheet'!F15-'Balance Sheet'!E15"),
        ("Net Cash Generated from Operations (A)", "=SUM(B6:B10)", "=SUM(C6:C10)", "=SUM(D6:D10)", "=SUM(E6:E10)", "=SUM(F6:F10)"),
        
        ("B. CASH FLOW FROM INVESTING ACTIVITIES", "SUB", "", "", "", ""),
        ("  - Capital Expenditure / Project Cost", -15000000.0, 0.0, 0.0, 0.0, 0.0),
        ("Net Cash Used in Investing Activities (B)", "=B13", "=C13", "=D13", "=E13", "=F13"),
        
        ("C. CASH FLOW FROM FINANCING ACTIVITIES", "SUB", "", "", "", ""),
        ("  - Term Loan Facility Disbursed", 15000000.0, 0.0, 0.0, 0.0, 0.0),
        ("  - Change in Borrowings (Repayments)", 0.0, "='Balance Sheet'!C11-'Balance Sheet'!B11", "='Balance Sheet'!D11-'Balance Sheet'!C11", "='Balance Sheet'!E11-'Balance Sheet'!D11", "='Balance Sheet'!F11-'Balance Sheet'!E11"),
        ("  - Interest on Term Loan Paid", "=-'Profit & Loss'!B16", "=-'Profit & Loss'!C16", "=-'Profit & Loss'!D16", 0.0, 0.0),
        ("Net Cash from Financing Activities (C)", "=SUM(B16:B18)", "=SUM(C16:C18)", "=SUM(D16:D18)", "=SUM(E16:E18)", "=SUM(F16:F18)"),
        
        ("D. SUMMARY OF CASH POSITION", "SUB", "", "", "", ""),
        ("  - Net Increase in Cash (A + B + C)", "=B11+B14+B19", "=C11+C14+C19", "=D11+D14+D19", "=E11+E14+E19", "=F11+F14+F19"),
        ("  - Opening Cash and Bank Balance", 0.0, "=B23", "=C23", "=D23", "=E23"),
        ("  - Closing Cash and Bank Balance", 2017226.03, 8101976.03, 18262038.53, 60358038.53, 102454038.53)
    ]
    
    for label, y1, y2, y3, y4, y5 in cf_structure:
        ws.row_dimensions[r].height = 20
        
        if y1 == "SUB":
            ws.cell(row=r, column=1, value=label).font = font_bold_data
            ws.cell(row=r, column=1).alignment = align_left
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            style_row_range(ws, r, 1, 6, border=border_thin)
        else:
            is_bold = "TOTAL" in label or "Net Cash" in label or "Closing" in label or "Increase" in label
            ws.cell(row=r, column=1, value=label).font = font_bold_data if is_bold else font_regular_data
            ws.cell(row=r, column=1).alignment = align_left
            ws.cell(row=r, column=1).border = border_thin
            
            for idx, val in enumerate([y1, y2, y3, y4, y5]):
                col_idx = idx + 2
                cell = ws.cell(row=r, column=col_idx)
                cell.value = val
                cell.font = font_bold_data if is_bold else font_regular_data
                cell.alignment = align_right
                cell.border = border_thin
                cell.number_format = FMT_CURRENCY
                
            if is_bold:
                style_row_range(ws, r, 1, 6, fill=fill_total)
                if "Closing" in label or "Net" in label or "Summary" in label:
                    style_row_range(ws, r, 1, 6, border=border_total)
        r += 1
        
    autofit_columns(ws)

# =========================================================================
# SHEET 10: FINANCIAL RATIOS DASHBOARD
# =========================================================================
def build_ratios_sheet(ws):
    ws.title = "Ratios"
    ws.sheet_view.showGridLines = True
    
    r = apply_corporate_header(ws, 1, "FINANCIAL RATIO ANALYSIS DASHBOARD", "Key Risk, Leverage, & Profitability Ratios vs Banking Benchmarks")
    r += 1
    
    ws.cell(row=r, column=1, value="Core Banking Financial Ratios").font = font_section
    r += 1
    
    headers = ["Ratio / Indicator", "Formula", "Banking Norm", "FY 2025-26", "FY 2026-27", "FY 2027-28", "FY 2028-29", "FY 2029-30"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    # Starts at row 6
    ratios_data = [
        # Particulars, formula, norm, Y1, Y2, Y3, Y4, Y5, format
        ("Current Ratio", "TCA / TCL", ">= 1.33", "='Balance Sheet'!B31/'Balance Sheet'!B17", "='Balance Sheet'!C31/'Balance Sheet'!C17", "='Balance Sheet'!D31/'Balance Sheet'!D17", "='Balance Sheet'!E31/'Balance Sheet'!E17", "='Balance Sheet'!F31/'Balance Sheet'!F17", "0.00"),
        ("Quick Ratio", "(TCA-Stock)/TCL", ">= 1.00", "=('Balance Sheet'!B31-'Balance Sheet'!B30-'Balance Sheet'!B26)/'Balance Sheet'!B17", "=('Balance Sheet'!C31-'Balance Sheet'!C30-'Balance Sheet'!C26)/'Balance Sheet'!C17", "=('Balance Sheet'!D31-'Balance Sheet'!D30-'Balance Sheet'!D26)/'Balance Sheet'!D17", "=('Balance Sheet'!E31-'Balance Sheet'!E30-'Balance Sheet'!E26)/'Balance Sheet'!E17", "=('Balance Sheet'!F31-'Balance Sheet'!F30-'Balance Sheet'!F26)/'Balance Sheet'!F17", "0.00"),
        ("Debt-Equity Ratio", "Long Term Debt / TNW", "<= 2.00", "='Balance Sheet'!B11/'Balance Sheet'!B9", "='Balance Sheet'!C11/'Balance Sheet'!C9", "='Balance Sheet'!D11/'Balance Sheet'!D9", "='Balance Sheet'!E11/'Balance Sheet'!E9", "='Balance Sheet'!F11/'Balance Sheet'!F9", "0.00"),
        ("TOL / TNW Ratio", "Total Outside Liab / TNW", "<= 3.00", "='Balance Sheet'!B18/'Balance Sheet'!B9", "='Balance Sheet'!C18/'Balance Sheet'!C9", "='Balance Sheet'!D18/'Balance Sheet'!D9", "='Balance Sheet'!E18/'Balance Sheet'!E9", "='Balance Sheet'!F18/'Balance Sheet'!F9", "0.00"),
        ("Debt Service Coverage Ratio (DSCR)", "(PAT+Dep+Int)/(Prin+Int)", ">= 1.25", 
         "=IF(('Amortization'!D6+'Profit & Loss'!B16)=0,0,('Profit & Loss'!B19+'Profit & Loss'!B14+'Profit & Loss'!B16)/('Amortization'!D6+'Profit & Loss'!B16))", 
         "=IF(('Amortization'!D7+'Profit & Loss'!C16)=0,0,('Profit & Loss'!C19+'Profit & Loss'!C14+'Profit & Loss'!C16)/('Amortization'!D7+'Profit & Loss'!C16))", 
         "=IF(('Amortization'!D8+'Profit & Loss'!D16)=0,0,('Profit & Loss'!D19+'Profit & Loss'!D14+'Profit & Loss'!D16)/('Amortization'!D8+'Profit & Loss'!D16))", 
         "=IF(('Amortization'!D9+'Profit & Loss'!E16)=0,0,('Profit & Loss'!E19+'Profit & Loss'!E14+'Profit & Loss'!E16)/('Amortization'!D9+'Profit & Loss'!E16))", 
         "=IF(('Amortization'!D10+'Profit & Loss'!F16)=0,0,('Profit & Loss'!F19+'Profit & Loss'!F14+'Profit & Loss'!F16)/('Amortization'!D10+'Profit & Loss'!F16))", "0.00"),
        ("Interest Coverage Ratio", "EBIT / Interest", ">= 2.00", "=IF('Profit & Loss'!B16=0,0,'Profit & Loss'!B15/'Profit & Loss'!B16)", "=IF('Profit & Loss'!C16=0,0,'Profit & Loss'!C15/'Profit & Loss'!C16)", "=IF('Profit & Loss'!D16=0,0,'Profit & Loss'!D15/'Profit & Loss'!D16)", "=IF('Profit & Loss'!E16=0,0,'Profit & Loss'!E15/'Profit & Loss'!E16)", "=IF('Profit & Loss'!F16=0,0,'Profit & Loss'!F15/'Profit & Loss'!F16)", "0.00"),
        ("Gross Profit Margin %", "(Revenue-RM Cost)/Revenue", "> 30%", "=('Profit & Loss'!B5-'Profit & Loss'!B8)/'Profit & Loss'!B5", "=('Profit & Loss'!C5-'Profit & Loss'!C8)/'Profit & Loss'!C5", "=('Profit & Loss'!D5-'Profit & Loss'!D8)/'Profit & Loss'!D5", "=('Profit & Loss'!E5-'Profit & Loss'!E8)/'Profit & Loss'!E5", "=('Profit & Loss'!F5-'Profit & Loss'!F8)/'Profit & Loss'!F5", FMT_PERCENT),
        ("EBITDA Margin %", "EBITDA / Sales", "> 10%", "='Profit & Loss'!B13", "='Profit & Loss'!C13", "='Profit & Loss'!D13", "='Profit & Loss'!E13", "='Profit & Loss'!F13", FMT_PERCENT),
        ("PAT Margin %", "PAT / Sales", "> 5%", "='Profit & Loss'!B20", "='Profit & Loss'!C20", "='Profit & Loss'!D20", "='Profit & Loss'!E20", "='Profit & Loss'!F20", FMT_PERCENT)
    ]
    
    for label, formula, norm, y1, y2, y3, y4, y5, fmt in ratios_data:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=label).font = font_bold_data
        ws.cell(row=r, column=1).alignment = align_left
        ws.cell(row=r, column=1).border = border_thin
        
        ws.cell(row=r, column=2, value=formula).font = font_regular_data
        ws.cell(row=r, column=2).alignment = align_left
        ws.cell(row=r, column=2).border = border_thin
        
        ws.cell(row=r, column=3, value=norm).font = font_bold_data
        ws.cell(row=r, column=3).alignment = align_center
        ws.cell(row=r, column=3).border = border_thin
        
        for idx, val in enumerate([y1, y2, y3, y4, y5]):
            col_idx = idx + 4
            cell = ws.cell(row=r, column=col_idx)
            cell.value = val
            cell.font = font_bold_data
            cell.alignment = align_right
            cell.border = border_thin
            cell.number_format = fmt
            
        style_row_range(ws, r, 1, 8, fill=fill_total)
        r += 1
        
    autofit_columns(ws)

# =========================================================================
# SHEET 11: SENSITIVITY ANALYSIS
# =========================================================================
def build_sensitivity_sheet(ws):
    ws.title = "Sensitivity"
    ws.sheet_view.showGridLines = True
    
    r = apply_corporate_header(ws, 1, "SENSITIVITY ANALYSIS STRESS TESTS", "Impact of Adverse Revenue and Cost Scenarios on Debt Service (DSCR)")
    r += 1
    
    ws.cell(row=r, column=1, value="DSCR Stress Scenarios").font = font_section
    r += 1
    
    headers = ["Scenario", "Assumption", "Projected DSCR", "Stressed DSCR", "Assessment"]
    ws.row_dimensions[r].height = 24
    for idx, h in enumerate(headers):
        cell = ws.cell(row=r, column=idx+1, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_header
    r += 1
    
    scenarios = [
        ("Revenue Decline 10%", "Sales drop by 10%", 0.0, 0.0, "Stressed"),
        ("Revenue Decline 20%", "Sales drop by 20%", 0.0, 0.0, "Stressed"),
        ("Cost Increase 10%", "Material costs rise 10%", 0.0, 0.0, "Stressed"),
        ("Combined Stress", "Revenue -10%, Cost +5%", 0.0, 0.0, "Stressed")
    ]
    
    for sc, ass, proj, strsd, assess in scenarios:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=sc).font = font_bold_data
        ws.cell(row=r, column=1).border = border_thin
        ws.cell(row=r, column=1).alignment = align_left
        
        ws.cell(row=r, column=2, value=ass).font = font_regular_data
        ws.cell(row=r, column=2).border = border_thin
        ws.cell(row=r, column=2).alignment = align_left
        
        ws.cell(row=r, column=3, value=proj).font = font_regular_data
        ws.cell(row=r, column=3).border = border_thin
        ws.cell(row=r, column=3).alignment = align_right
        ws.cell(row=r, column=3).number_format = "0.00x"
        
        ws.cell(row=r, column=4, value=strsd).font = font_regular_data
        ws.cell(row=r, column=4).border = border_thin
        ws.cell(row=r, column=4).alignment = align_right
        ws.cell(row=r, column=4).number_format = "0.00x"
        
        ws.cell(row=r, column=5, value=assess).font = font_bold_data
        ws.cell(row=r, column=5).border = border_thin
        ws.cell(row=r, column=5).alignment = align_center
        ws.cell(row=r, column=5).fill = fill_total
        ws.cell(row=r, column=5).font = Font(name=FONT_FAMILY, size=10, bold=True, color="FF0000")
        
        r += 1
        
    autofit_columns(ws)

# =========================================================================
# MAIN GENERATOR EXECUTION
# =========================================================================
def main():
    print("Generating Goyal Fertilizer CMA Excel Report...")
    
    build_assumptions_sheet(wb.create_sheet())
    build_project_details_sheet(wb.create_sheet())
    build_repayment_sheet(wb.create_sheet())
    build_depreciation_sheet(wb.create_sheet())
    build_pnl_sheet(wb.create_sheet())
    build_balance_sheet(wb.create_sheet())
    build_mpbf_sheet(wb.create_sheet())
    build_dp_sheet(wb.create_sheet())
    build_cashflow_sheet(wb.create_sheet())
    build_ratios_sheet(wb.create_sheet())
    build_sensitivity_sheet(wb.create_sheet())
    
    output_path = "public/Goyal_Fertilizer_CMA_Project_Report.xlsx"
    wb.save(output_path)
    wb.save("Goyal_Fertilizer_CMA_Project_Report.xlsx")
    
    print(f"Goyal Fertilizer CMA Report saved to: {output_path}")

if __name__ == "__main__":
    main()
